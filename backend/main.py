if True:
    from datetime import datetime
    from fastapi import BackgroundTasks, FastAPI, UploadFile, File, Form, Depends, HTTPException, Body
    from fastapi.middleware.cors import CORSMiddleware
    from sqlalchemy.orm import Session
    from sqlalchemy.exc import IntegrityError
    from sqlalchemy import func, distinct, inspect, update, create_engine, text, or_, and_, exists
    from pydantic import BaseModel
    from typing import List, Optional
    import shutil
    import os
    import json
    import time
    import psycopg2
    from psycopg2 import sql as psql
    import pandas as pd
    from urllib.parse import quote_plus
    from urllib import request as urlrequest
    from urllib import parse as urlparse
    import base64
    import atexit
    import sys
    from settings import get_cors_origins, load_env

    load_env()
    from database import POSTGRES_DB, SessionLocal, engine

    import models  # noqa: E402 — must run after database.py configures the engine

    os.makedirs("uploads", exist_ok=True)
    from utils.upload_paths import (
        job_temp_upload_path,
        rename_table_csv,
        resolve_table_csv_path,
        table_csv_path,
    )
    from engine.orchestrator import run_data_quality_job
    from fastapi.responses import StreamingResponse
    import io
    import zipfile
    from openpyxl.styles import PatternFill
    from openpyxl import load_workbook
    from thefuzz import process
    from pydantic import BaseModel

    from prometheus_fastapi_instrumentator import Instrumentator
    from apscheduler.schedulers.background import BackgroundScheduler
    from apscheduler.triggers.cron import CronTrigger
    from starlette.requests import Request

    from auth.access_routes import router as access_router
    from auth.admin_routes import router as admin_router
    from auth.middleware import auth_middleware
    from auth.routes import router as auth_router
    from routers.admin.router import router as platform_admin_router
    from routers.audit.router import router as audit_router
    from routers.compliance.router import router as compliance_router
    from routers.dashboard.router import router as dashboard_router
    from routers.governance.router import router as governance_router
    from routers.lineage.router import router as lineage_router
    from routers.reports.router import router as reports_router
    from routers.stewardship.router import router as stewardship_router
    from routers.enterprise.router import router as enterprise_router
    from auth.deps import get_current_user
    app = FastAPI(title="Data Quality Engine")


    # Create required schemas + tables
    def _init_database_schema():
        from database import POSTGRES_HOST, POSTGRES_DB

        print(
            f"[mdqm] Initializing schemas on {POSTGRES_HOST}/{POSTGRES_DB}...",
            file=sys.stderr,
            flush=True,
        )
        with engine.begin() as conn:
            conn.execute(text("CREATE SCHEMA IF NOT EXISTS auth"))
            conn.execute(text("CREATE SCHEMA IF NOT EXISTS governance"))
            conn.execute(text("CREATE SCHEMA IF NOT EXISTS metadata"))
            conn.execute(text("CREATE SCHEMA IF NOT EXISTS quarantine"))
            conn.execute(text("CREATE SCHEMA IF NOT EXISTS enterprise"))
        models.Base.metadata.create_all(bind=engine)
        # SQLAlchemy create_all does not alter existing tables; add column if model gained it via migration.
        with engine.begin() as conn:
            conn.execute(
                text(
                    "ALTER TABLE metadata.jobs ADD COLUMN IF NOT EXISTS db_source_config JSONB"
                )
            )
            conn.execute(
                text(
                    "ALTER TABLE metadata.db_connections ADD COLUMN IF NOT EXISTS user_id INTEGER"
                )
            )
            conn.execute(
                text(
                    "ALTER TABLE metadata.db_connections ADD COLUMN IF NOT EXISTS db_type VARCHAR(32) DEFAULT 'postgres'"
                )
            )
        print("[mdqm] Database schema ready.", file=sys.stderr, flush=True)

    def _seed_users():
        from auth.seed import seed_users_on_startup

        db = SessionLocal()
        try:
            seed_users_on_startup(db)
        finally:
            db.close()

    try:
        _init_database_schema()
        _seed_users()
    except Exception as exc:
        print(f"[mdqm] FATAL: Database startup failed: {exc}", file=sys.stderr, flush=True)
        raise

    try:
        _access_ddl = [
            "ALTER TABLE auth.access_requests ADD COLUMN IF NOT EXISTS username VARCHAR(64)",
            "ALTER TABLE auth.access_requests ADD COLUMN IF NOT EXISTS dataset_name VARCHAR(255)",
            "ALTER TABLE auth.access_requests ADD COLUMN IF NOT EXISTS access_type VARCHAR(32)",
            "ALTER TABLE auth.access_requests ADD COLUMN IF NOT EXISTS duration VARCHAR(64)",
            "ALTER TABLE auth.access_requests ADD COLUMN IF NOT EXISTS approver_name VARCHAR(255)",
        ]
        for ddl in _access_ddl:
            with engine.begin() as conn:
                conn.execute(text(ddl))
    except Exception:
        pass

    # --- 1. CORS (local dev + GitHub Pages via CORS_ORIGINS env) ---
    app.add_middleware(
        CORSMiddleware,
        allow_origins=get_cors_origins(),
        allow_credentials=True,
        allow_methods=["*"],
        allow_headers=["*"],
        expose_headers=["Content-Disposition", "X-Export-Id"],
    )

    app.include_router(auth_router)
    app.include_router(access_router)
    app.include_router(admin_router)
    app.include_router(dashboard_router)
    app.include_router(governance_router)
    app.include_router(lineage_router)
    app.include_router(compliance_router)
    app.include_router(audit_router)
    app.include_router(stewardship_router)
    app.include_router(reports_router)
    app.include_router(platform_admin_router)
    app.include_router(enterprise_router)


    @app.middleware("http")
    async def enterprise_api_logging(request: Request, call_next):
        if not str(request.url.path).startswith("/api/enterprise"):
            return await call_next(request)
        start = time.perf_counter()
        response = await call_next(request)
        duration_ms = int((time.perf_counter() - start) * 1000)
        db = None
        try:
            db = SessionLocal()
            db.add(
                models.EnterpriseApiLog(
                    method=request.method,
                    path=str(request.url.path)[:512],
                    status_code=getattr(response, "status_code", None),
                    duration_ms=duration_ms,
                    user_id=getattr(request.state, "user_id", None),
                    correlation_id=getattr(request.state, "correlation_id", None),
                    ip_address=request.client.host if request.client else None,
                )
            )
            db.commit()
        except Exception:
            if db:
                try:
                    db.rollback()
                except Exception:
                    pass
        finally:
            if db:
                try:
                    db.close()
                except Exception:
                    pass
        return response

    @app.middleware("http")
    async def rbac_middleware(request: Request, call_next):
        return await auth_middleware(request, call_next)

    scheduler = BackgroundScheduler()
    scheduler.start()
    atexit.register(lambda: scheduler.shutdown(wait=False))


    def _run_scheduled_job(job_id: int):
        db = SessionLocal()
        try:
            run_data_quality_job(job_id, db)
        finally:
            db.close()

    def _run_scheduled_refresh(job_id: int):
        _execute_import_for_job(job_id, body=None)


    def _scheduler_job_key(job_id: int) -> str:
        return f"scheduled_job_{job_id}"

    def _scheduler_refresh_job_key(job_id: int) -> str:
        return f"{_scheduler_job_key(job_id)}_refresh"

    def _parse_scheduled_job_id(scheduler_id: str) -> int | None:
        raw = str(scheduler_id or "")
        if not raw.startswith("scheduled_job_"):
            return None
        suffix = raw.replace("scheduled_job_", "", 1)
        if suffix.endswith("_refresh"):
            suffix = suffix[: -len("_refresh")]
        try:
            return int(suffix)
        except ValueError:
            return None

    def _schedule_action_from_id(scheduler_id: str) -> str:
        return "refresh" if str(scheduler_id or "").endswith("_refresh") else "dq"

    def _find_scheduled_job(job_id: int, *, prefer_refresh: bool = False):
        keys = (
            [_scheduler_refresh_job_key(job_id), _scheduler_job_key(job_id)]
            if prefer_refresh
            else [_scheduler_job_key(job_id), _scheduler_refresh_job_key(job_id)]
        )
        for key in keys:
            j = scheduler.get_job(key)
            if j:
                return j, key
        return None, None

    def _serialize_schedule_job(job):
        sid = str(getattr(job, "id", ""))
        return {
            "scheduler_id": sid,
            "job_id": _parse_scheduled_job_id(sid),
            "action": _schedule_action_from_id(sid),
            "next_run_time": job.next_run_time.isoformat() if getattr(job, "next_run_time", None) else None,
            "trigger": str(getattr(job, "trigger", "")),
            "paused": bool(getattr(job, "next_run_time", None) is None),
        }

    def get_db():
        db = SessionLocal()
        try:
            yield db
        finally:
            db.close()

    # --- PYDANTIC SCHEMAS ---
    class RuleCreate(BaseModel):
        job_id: int
        table_id: int
        column_name: str
        rule_type: str
        data_type: str
        rule_value: Optional[str] = None
        is_active: bool = True
        master_data: Optional[List[str]] = []

    class RuleToggle(BaseModel):
        is_active: bool

    class RuleUpdate(BaseModel):
        rule_type: str
        rule_value: Optional[str] = None
        is_active: bool
        master_data: Optional[List[str]] = [] # For updating fuzzy lists
        
    class RenamePayload(BaseModel):
        name: str
        
    class ErrorEdit(BaseModel):
        new_value: str
        
    class MasterAdd(BaseModel):
        new_master: str

    class FuzzyReplace(BaseModel):
        row_id: int
        column_name: str
        new_value: str
        
    class JobCreate(BaseModel):
        job_name: str


    class TableEmailPayload(BaseModel):
        to_email: str
        format: Optional[str] = "csv"
        subject: Optional[str] = None
        body: Optional[str] = None

    # Legacy file for saved DB connections (migrated once into metadata.db_connections).
    CONNECTIONS_FILE = os.path.join(os.path.dirname(__file__), "saved_connections.json")
    SOURCE_PATHS_FILE = os.path.join(os.path.dirname(__file__), "source_paths.json")


    def _sync_db_connections_id_sequence(db: Session) -> None:
        """Keep PostgreSQL sequence past MAX(connection id) after explicit inserts."""
        try:
            db.execute(
                text(
                    "SELECT setval(pg_get_serial_sequence('metadata.db_connections', 'connection_id'), "
                    "GREATEST(COALESCE((SELECT MAX(connection_id) FROM metadata.db_connections), 1), 1))"
                )
            )
            db.commit()
        except Exception:
            db.rollback()


    def _migrate_saved_connections_json_to_db(db: Session) -> None:
        if not os.path.isfile(CONNECTIONS_FILE):
            return
        try:
            with open(CONNECTIONS_FILE, "r", encoding="utf-8") as f:
                data = json.load(f)
        except Exception:
            return
        items = data if isinstance(data, list) else []
        from utils.source_secret_crypto import encrypt_db_password_optional, encryption_available

        migrated_any = False
        for item in items:
            if not isinstance(item, dict):
                continue
            try:
                cid = int(float(item.get("connection_id", 0)))
            except (TypeError, ValueError):
                continue
            if cid <= 0:
                continue
            if db.query(models.DbConnection).filter(models.DbConnection.connection_id == cid).first():
                continue
            name = str(item.get("connection_name") or "").strip() or f"connection_{cid}"
            if db.query(models.DbConnection).filter(models.DbConnection.connection_name == name).first():
                continue
            raw_pass = str(item.get("pass", "") or "")
            stored_pw = None
            if raw_pass.strip():
                if encryption_available():
                    stored_pw = encrypt_db_password_optional(raw_pass) or raw_pass
                else:
                    stored_pw = raw_pass
            row = models.DbConnection(
                connection_id=cid,
                connection_name=name,
                host=str(item.get("host") or ""),
                port=str(item.get("port") or "5432"),
                username=str(item.get("user") or "").strip(),
                password=stored_pw,
            )
            db.add(row)
            migrated_any = True
        try:
            if migrated_any:
                db.commit()
                _sync_db_connections_id_sequence(db)
        except Exception:
            db.rollback()
            return
        try:
            os.replace(CONNECTIONS_FILE, CONNECTIONS_FILE + ".migrated")
        except OSError:
            pass


    def _stored_connection_password_plain(row: models.DbConnection) -> str:
        """Decrypt saved password or return legacy plaintext."""
        from utils.source_secret_crypto import decrypt_db_password_optional

        blob = row.password
        if not blob or not str(blob).strip():
            return ""
        plain = decrypt_db_password_optional(blob)
        if plain is not None:
            return plain
        return str(blob)


    def _read_source_paths():
        if not os.path.exists(SOURCE_PATHS_FILE):
            return {}
        try:
            with open(SOURCE_PATHS_FILE, "r", encoding="utf-8") as f:
                data = json.load(f)
                return data if isinstance(data, dict) else {}
        except Exception:
            return {}


    def _write_source_paths(items: dict):
        with open(SOURCE_PATHS_FILE, "w", encoding="utf-8") as f:
            json.dump(items, f, ensure_ascii=True, indent=2)


    def _source_path_key(job_id: int, table_id: int) -> str:
        return f"{job_id}:{table_id}"


    def _save_table_source_path(job_id: int, table_id: int, source_path: str):
        items = _read_source_paths()
        items[_source_path_key(job_id, table_id)] = _normalize_local_path(source_path)
        _write_source_paths(items)


    def _get_table_source_path(job_id: int, table_id: int):
        items = _read_source_paths()
        return items.get(_source_path_key(job_id, table_id))


    def _refresh_table_csv_from_source_path(job_id: int, table_id: int, db: Session):
        """
        If a source file path was configured for this table, reload that source into uploads/<table>.csv
        so each Run Job uses current file contents.
        """
        source_path = _get_table_source_path(job_id, table_id)
        if not source_path:
            return
        if not os.path.isfile(source_path):
            return

        table = (
            db.query(models.TableMetadata)
            .filter(models.TableMetadata.job_id == job_id, models.TableMetadata.table_id == table_id)
            .first()
        )
        if not table:
            return

        final_csv_path = table_csv_path(job_id, table.table_name)

        if source_path.lower().endswith((".xlsx", ".xls")):
            df = pd.read_excel(source_path)
        else:
            df = pd.read_csv(source_path)
        df.to_csv(final_csv_path, index=False)

        # Keep metadata aligned with latest source snapshot.
        table.row_count = int(len(df))
        db.query(models.ColumnMetadata).filter(
            models.ColumnMetadata.job_id == job_id,
            models.ColumnMetadata.table_id == table_id,
        ).delete(synchronize_session=False)

        for col_name, dtype in df.dtypes.items():
            str_type = "String"
            if "int" in str(dtype):
                str_type = "Integer"
            elif "float" in str(dtype):
                str_type = "Float"
            elif "datetime" in str(dtype):
                str_type = "Date"
            elif "bool" in str(dtype):
                str_type = "Boolean"
            db.add(
                models.ColumnMetadata(
                    job_id=job_id,
                    table_id=table_id,
                    column_name=col_name,
                    data_type=str_type,
                )
            )

        # Clear old stats so pre-run UI does not show stale totals.
        db.query(models.TableStats).filter(
            models.TableStats.job_id == job_id,
            models.TableStats.table_id == table_id,
        ).delete(synchronize_session=False)

        db.commit()


    def _normalize_local_path(raw_path: str):
        """
        Normalize user-entered Windows paths.
        Accepts sloppy forms like 'c\\downloads\\file.csv' by converting to 'c:\\downloads\\file.csv'.
        """
        value = (raw_path or "").strip().strip('"').strip("'")
        if not value:
            return value

        # Handle drive typed as "c\foo\bar.csv" (missing colon)
        if len(value) >= 2 and value[1] == "\\" and value[0].isalpha():
            value = f"{value[0]}:{value[1:]}"

        # Normalize slashes and relative components.
        return os.path.normpath(value)


    def _excel_preview_fast(file_path: str):
        """
        Fast preview for large XLSX files without loading whole sheet into pandas.
        Returns columns, column_types, rows(first 11), total_rows.
        """
        wb = load_workbook(file_path, read_only=True, data_only=True)
        ws = wb.active

        row_iter = ws.iter_rows(values_only=True)
        header = next(row_iter, None)
        if not header:
            wb.close()
            return {"columns": [], "column_types": {}, "rows": [], "total_rows": 0}

        columns = [str(c).strip() if c is not None else "" for c in header]
        columns = [c if c else f"col_{idx+1}" for idx, c in enumerate(columns)]

        sample_rows = []
        type_samples = {c: [] for c in columns}
        for idx, row in enumerate(row_iter, start=1):
            row_vals = list(row[: len(columns)])
            if idx <= 11:
                sample_rows.append({col: ("" if val is None else val) for col, val in zip(columns, row_vals)})
            if idx <= 500:
                for col, val in zip(columns, row_vals):
                    if val is not None and val != "":
                        type_samples[col].append(val)
            if idx >= 2000:
                break

        def infer(vals):
            if not vals:
                return "string"
            if all(isinstance(v, bool) for v in vals):
                return "bool"
            # bool is a subclass of int, so this check must be after bool.
            if all(isinstance(v, int) and not isinstance(v, bool) for v in vals):
                return "int64"
            if all(isinstance(v, (int, float)) and not isinstance(v, bool) for v in vals):
                return "float64"
            if all(isinstance(v, datetime) for v in vals):
                return "datetime64[ns]"
            return "object"

        column_types = {col: infer(type_samples[col]) for col in columns}
        total_rows = max((ws.max_row or 1) - 1, 0)
        wb.close()

        return {
            "columns": columns,
            "column_types": column_types,
            "rows": sample_rows,
            "total_rows": int(total_rows),
        }
    # --- 3. API ENDPOINTS ---

    @app.get("/")
    def read_root():
        return {"message": "MDQM Backend is Live"}

    @app.get("/health")
    def health_check():
        return {"status": "ok"}


    @app.post("/files/preview")
    async def preview_file(file: UploadFile = File(...)):
        """Return first rows + inferred column types for CSV / Excel."""
        try:
            if file.filename.lower().endswith((".xlsx", ".xls")):
                df = pd.read_excel(file.file)
            else:
                df = pd.read_csv(file.file)
        except Exception as e:
            raise HTTPException(status_code=400, detail=f"Invalid file: {str(e)}")

        rows = df.head(11).fillna("").to_dict(orient="records")
        column_types = {col: str(dtype) for col, dtype in df.dtypes.items()}
        return {
            "columns": df.columns.tolist(),
            "column_types": column_types,
            "rows": rows,
            "total_rows": int(len(df)),
        }


    @app.post("/files/preview-from-path")
    def preview_file_from_path(payload: dict = Body(...)):
        file_path = _normalize_local_path(payload.get("file_path"))
        if not file_path:
            raise HTTPException(status_code=400, detail="file_path is required")
        if not os.path.isfile(file_path):
            raise HTTPException(
                status_code=400,
                detail=f"File path not found. Use a full path like C:\\Downloads\\data.csv. Received: {file_path}",
            )

        try:
            lower = file_path.lower()
            if lower.endswith((".xlsx", ".xls")):
                return _excel_preview_fast(file_path)
            else:
                df = pd.read_csv(file_path)
        except Exception as e:
            raise HTTPException(status_code=400, detail=f"Invalid file: {str(e)}")

        rows = df.head(11).fillna("").to_dict(orient="records")
        column_types = {col: str(dtype) for col, dtype in df.dtypes.items()}
        return {
            "columns": df.columns.tolist(),
            "column_types": column_types,
            "rows": rows,
            "total_rows": int(len(df)),
        }


    @app.get("/db/connections")
    def list_db_connections(db: Session = Depends(get_db), user: models.User = Depends(get_current_user)):
        _migrate_saved_connections_json_to_db(db)
        shared_exists = (
            db.query(models.DbConnectionShare)
            .filter(
                models.DbConnectionShare.connection_id == models.DbConnection.connection_id,
                models.DbConnectionShare.shared_user_id == user.id,
            )
            .exists()
        )
        rows = (
            db.query(models.DbConnection)
            .filter(
                or_(
                    models.DbConnection.user_id == user.id,
                    models.DbConnection.user_id.is_(None),
                    shared_exists,
                )
            )
            .order_by(models.DbConnection.connection_id.asc())
            .all()
        )
        return [
            {
                "connection_id": r.connection_id,
                "connection_name": r.connection_name or "",
                "host": r.host or "",
                "port": r.port or "5432",
                "user": r.username or "",
                "db_type": r.db_type or "postgres",
                "owned": bool(r.user_id == user.id),
            }
            for r in rows
        ]


    @app.get("/db/connections/{connection_id}/credentials")
    def get_saved_connection_credentials(
        connection_id: int,
        db: Session = Depends(get_db),
        user: models.User = Depends(get_current_user),
    ):
        """
        Return decrypted credentials for a saved profile so the UI can pre-fill host/port/user/password.
        Requires authentication (same as other /db routes). Password is at rest encrypted in metadata.db_connections.
        """
        _migrate_saved_connections_json_to_db(db)
        shared_exists = (
            db.query(models.DbConnectionShare)
            .filter(
                models.DbConnectionShare.connection_id == models.DbConnection.connection_id,
                models.DbConnectionShare.shared_user_id == user.id,
            )
            .exists()
        )
        row = (
            db.query(models.DbConnection)
            .filter(models.DbConnection.connection_id == connection_id)
            .filter(
                or_(
                    models.DbConnection.user_id == user.id,
                    models.DbConnection.user_id.is_(None),
                    shared_exists,
                )
            )
            .first()
        )
        if not row:
            raise HTTPException(status_code=404, detail="Saved connection not found")
        return {
            "connection_id": row.connection_id,
            "connection_name": row.connection_name or "",
            "host": row.host or "",
            "port": row.port or "5432",
            "user": row.username or "",
            "password": _stored_connection_password_plain(row),
            "db_type": row.db_type or "postgres",
        }


    @app.post("/db/connections")
    def save_db_connection(
        payload: dict = Body(...),
        db: Session = Depends(get_db),
        user: models.User = Depends(get_current_user),
    ):
        _migrate_saved_connections_json_to_db(db)
        resolved_user = payload.get("user", payload.get("username", ""))
        resolved_pass = payload.get("pass", payload.get("password", ""))
        plain_pw = str(resolved_pass or "").strip()
        if plain_pw:
            from utils.source_secret_crypto import encrypt_db_password_optional, encryption_available

            if not encryption_available():
                raise HTTPException(
                    status_code=400,
                    detail=(
                        "Set MDQM_DB_SOURCE_MASTER_SECRET in the backend .env (see .env.example). "
                        "It is required so saved connection passwords are encrypted in the database."
                    ),
                )
            enc = encrypt_db_password_optional(plain_pw)
            if not enc:
                raise HTTPException(
                    status_code=400,
                    detail=(
                        "Password encryption failed. Install the cryptography package in the backend "
                        "(e.g. pip install cryptography), then restart the API."
                    ),
                )
            stored_pw = enc
        else:
            stored_pw = None
        name = str(payload.get("connection_name") or "").strip()
        if not name:
            raise HTTPException(status_code=400, detail="connection_name is required.")
        row = models.DbConnection(
            connection_name=name,
            host=str(payload.get("host") or ""),
            port=str(payload.get("port") or "5432"),
            username=str(resolved_user or "").strip(),
            password=stored_pw,
            user_id=user.id,
            db_type=str(payload.get("db_type") or "postgres").strip().lower(),
        )
        db.add(row)
        try:
            db.commit()
            db.refresh(row)
        except IntegrityError:
            db.rollback()
            raise HTTPException(
                status_code=400,
                detail="A saved connection with this name already exists.",
            )
        return {"message": "Connection saved", "connection_id": row.connection_id}


    @app.put("/db/connections/{connection_id}")
    def update_db_connection(
        connection_id: int,
        payload: dict = Body(...),
        db: Session = Depends(get_db),
        user: models.User = Depends(get_current_user),
    ):
        _migrate_saved_connections_json_to_db(db)
        row = (
            db.query(models.DbConnection)
            .filter(models.DbConnection.connection_id == connection_id)
            .filter(models.DbConnection.user_id == user.id)
            .first()
        )
        if not row:
            raise HTTPException(status_code=404, detail="Saved connection not found")

        name = str(payload.get("connection_name") or "").strip()
        if not name:
            raise HTTPException(status_code=400, detail="connection_name is required.")

        row.connection_name = name
        row.host = str(payload.get("host") or "")
        row.port = str(payload.get("port") or "5432")
        row.username = str(payload.get("user") or "").strip()
        row.db_type = str(payload.get("db_type") or "postgres").strip().lower()

        if payload.get("pass") is not None:
            plain_pw = str(payload.get("pass") or "").strip()
            if plain_pw:
                from utils.source_secret_crypto import encrypt_db_password_optional, encryption_available
                if not encryption_available():
                    raise HTTPException(
                        status_code=400,
                        detail=(
                            "Set MDQM_DB_SOURCE_MASTER_SECRET in the backend .env (see .env.example). "
                            "It is required so saved connection passwords are encrypted in the database."
                        ),
                    )
                enc = encrypt_db_password_optional(plain_pw)
                if not enc:
                    raise HTTPException(
                        status_code=400,
                        detail=(
                            "Password encryption failed. Install the cryptography package in the backend "
                            "(e.g. pip install cryptography), then restart the API."
                        ),
                    )
                row.password = enc
            else:
                row.password = None

        try:
            db.commit()
        except IntegrityError:
            db.rollback()
            raise HTTPException(status_code=400, detail="A saved connection with this name already exists.")

        return {"message": "Connection updated", "connection_id": row.connection_id}


    @app.delete("/db/connections/{connection_id}")
    def delete_db_connection(
        connection_id: int,
        db: Session = Depends(get_db),
        user: models.User = Depends(get_current_user),
    ):
        row = (
            db.query(models.DbConnection)
            .filter(models.DbConnection.connection_id == connection_id)
            .filter(models.DbConnection.user_id == user.id)
            .first()
        )
        if not row:
            raise HTTPException(status_code=404, detail="Saved connection not found")
        db.delete(row)
        db.commit()
        return {"message": "Connection deleted"}


    @app.get("/db/connections/{connection_id}/shares")
    def list_connection_shares(
        connection_id: int,
        db: Session = Depends(get_db),
        user: models.User = Depends(get_current_user),
    ):
        _migrate_saved_connections_json_to_db(db)
        row = (
            db.query(models.DbConnection)
            .filter(models.DbConnection.connection_id == connection_id)
            .filter(models.DbConnection.user_id == user.id)
            .first()
        )
        if not row:
            raise HTTPException(status_code=404, detail="Saved connection not found")

        shares = (
            db.query(models.DbConnectionShare, models.User)
            .join(models.User, models.DbConnectionShare.shared_user_id == models.User.id)
            .filter(models.DbConnectionShare.connection_id == connection_id)
            .all()
        )
        return [
            {
                "share_id": share.share_id,
                "shared_user_id": user.id,
                "shared_username": user.username,
                "shared_email": user.email,
            }
            for share, user in shares
        ]


    @app.post("/db/connections/{connection_id}/share")
    def share_db_connection(
        connection_id: int,
        payload: dict = Body(...),
        db: Session = Depends(get_db),
        user: models.User = Depends(get_current_user),
    ):
        _migrate_saved_connections_json_to_db(db)
        row = (
            db.query(models.DbConnection)
            .filter(models.DbConnection.connection_id == connection_id)
            .filter(models.DbConnection.user_id == user.id)
            .first()
        )
        if not row:
            raise HTTPException(status_code=404, detail="Saved connection not found")

        share_with = str(payload.get("share_with") or "").strip()
        if not share_with:
            raise HTTPException(status_code=400, detail="share_with is required.")

        target_user = (
            db.query(models.User)
            .filter(
                or_(models.User.username == share_with, models.User.email == share_with)
            )
            .first()
        )
        if not target_user:
            raise HTTPException(status_code=404, detail="Target user not found.")
        if target_user.id == user.id:
            raise HTTPException(status_code=400, detail="Cannot share with yourself.")

        existing_share = (
            db.query(models.DbConnectionShare)
            .filter(
                models.DbConnectionShare.connection_id == connection_id,
                models.DbConnectionShare.shared_user_id == target_user.id,
            )
            .first()
        )
        if existing_share:
            return {"message": "Connection already shared with this user."}

        share_row = models.DbConnectionShare(
            connection_id=connection_id,
            shared_user_id=target_user.id,
        )
        db.add(share_row)
        db.commit()
        return {"message": f"Connection shared with {target_user.username or target_user.email}."}


    @app.delete("/db/connections/{connection_id}/share/{target_user_id}")
    def unshare_db_connection(
        connection_id: int,
        target_user_id: int,
        db: Session = Depends(get_db),
        user: models.User = Depends(get_current_user),
    ):
        row = (
            db.query(models.DbConnection)
            .filter(models.DbConnection.connection_id == connection_id)
            .filter(models.DbConnection.user_id == user.id)
            .first()
        )
        if not row:
            raise HTTPException(status_code=404, detail="Saved connection not found")

        share_row = (
            db.query(models.DbConnectionShare)
            .filter(models.DbConnectionShare.connection_id == connection_id)
            .filter(models.DbConnectionShare.shared_user_id == target_user_id)
            .first()
        )
        if not share_row:
            raise HTTPException(status_code=404, detail="Share not found")

        db.delete(share_row)
        db.commit()
        return {"message": "Connection access revoked."}


    def _resolve_connection_payload(payload: dict, current_user_id: int | None = None):
        """Resolve credentials from direct fields or saved connection id."""
        connection_id = payload.get("connection_id")
        if connection_id is not None:
            db = SessionLocal()
            try:
                _migrate_saved_connections_json_to_db(db)
                try:
                    cid = int(float(connection_id))
                except (TypeError, ValueError):
                    raise HTTPException(status_code=404, detail="Saved connection not found")

                query = db.query(models.DbConnection).filter(models.DbConnection.connection_id == cid)
                if current_user_id is not None:
                    shared_exists = (
                        db.query(models.DbConnectionShare)
                        .filter(
                            models.DbConnectionShare.connection_id == models.DbConnection.connection_id,
                            models.DbConnectionShare.shared_user_id == current_user_id,
                        )
                        .exists()
                    )
                    query = query.filter(
                        or_(models.DbConnection.user_id == current_user_id, models.DbConnection.user_id.is_(None), shared_exists)
                    )
                row = query.first()
                if not row:
                    raise HTTPException(status_code=404, detail="Saved connection not found")
                db_type = row.db_type or "postgres"
                port_str = str(row.port or "").strip()
                # Defensive auto-detection for older/unmodified connection profiles
                if db_type == "postgres":
                    if port_str == "1433" or "sql server" in (row.connection_name or "").lower() or "mssql" in (row.connection_name or "").lower():
                        db_type = "sqlserver"
                    elif port_str == "3306" or "mysql" in (row.connection_name or "").lower():
                        db_type = "mysql"

                creds = {
                    "host": row.host,
                    "port": str(row.port or "5432"),
                    "user": row.username,
                    "pass": _stored_connection_password_plain(row),
                    "dbname": payload.get("dbname"),
                    "db_type": db_type,
                }
            finally:
                db.close()

            if str(payload.get("host") or "").strip():
                creds["host"] = str(payload["host"]).strip()
            if str(payload.get("port") or "").strip():
                creds["port"] = str(payload["port"]).strip()
            if str(payload.get("user") or "").strip():
                creds["user"] = str(payload.get("user")).strip()
            pw = payload.get("pass")
            if pw is not None and str(pw):
                creds["pass"] = str(pw)
            if payload.get("db_type"):
                creds["db_type"] = str(payload["db_type"]).strip().lower()
            return creds

        db_type = payload.get("db_type") or "postgres"
        port_str = str(payload.get("port") or "").strip()
        if db_type == "postgres":
            if port_str == "1433":
                db_type = "sqlserver"
            elif port_str == "3306":
                db_type = "mysql"

        return {
            "host": payload.get("host"),
            "port": str(payload.get("port", "5432")),
            "user": payload.get("user"),
            "pass": payload.get("pass", ""),
            "dbname": payload.get("dbname"),
            "db_type": db_type,
        }


    def _connect_postgres_with_fallback(creds: dict):
        """
        Connect using provided credentials first.
        If auth fails, retry once with backend .env credentials for local DB usage.
        """
        # Do not silently switch credentials/database. Users expect writes to happen
        # exactly in the DB they selected in UI.
        return psycopg2.connect(
            host=creds["host"],
            port=creds["port"],
            user=creds["user"],
            password=creds["pass"],
            dbname=creds["dbname"],
            connect_timeout=8,
        )

    def _connect_db(creds: dict):
        db_type = str(creds.get("db_type") or "postgres").lower().strip()
        if db_type in ("mssql", "sqlserver", "sql_server"):
            import pyodbc
            drivers = [
                "ODBC Driver 18 for SQL Server",
                "ODBC Driver 17 for SQL Server",
                "ODBC Driver 13 for SQL Server",
                "SQL Server Native Client 11.0",
                "SQL Server",
            ]
            last_err = None
            for driver in drivers:
                try:
                    host = creds["host"]
                    port = str(creds.get("port") or "").strip()
                    host_port = f"{host},{port}" if port and port not in ("0", "5432") else host
                    conn_str = (
                        f"Driver={{{driver}}};Server={host_port};"
                        f"Database={creds['dbname']};Uid={creds['user']};Pwd={creds['pass']};"
                        "TrustServerCertificate=yes;Connection Timeout=8;"
                    )
                    return pyodbc.connect(conn_str)
                except Exception as e:
                    last_err = e
            raise Exception(f"SQL Server connection failed: {last_err}")
        elif db_type == "mysql":
            import pymysql
            port = int(creds.get("port") or 3306)
            return pymysql.connect(
                host=creds["host"],
                port=port,
                user=creds["user"],
                password=creds["pass"],
                database=creds["dbname"],
                connect_timeout=8,
            )
        elif db_type == "oracle":
            import oracledb
            return oracledb.connect(
                user=creds["user"],
                password=creds["pass"],
                host=creds["host"],
                port=int(creds.get("port") or 1521),
                service_name=creds["dbname"],
            )
        elif db_type == "snowflake":
            import snowflake.connector
            return snowflake.connector.connect(
                user=creds["user"],
                password=creds["pass"],
                account=creds["host"],
                database=creds["dbname"],
            )
        elif db_type == "databricks":
            from databricks import sql
            server_hostname = creds["host"]
            http_path = creds["dbname"]
            if "/" in server_hostname:
                parts = server_hostname.split("/", 1)
                server_hostname = parts[0]
                http_path = "/" + parts[1]
            return sql.connect(
                server_hostname=server_hostname,
                http_path=http_path,
                access_token=creds["pass"],
            )
        else:
            return psycopg2.connect(
                host=creds["host"],
                port=creds["port"],
                user=creds["user"],
                password=creds["pass"],
                dbname=creds["dbname"],
                connect_timeout=8,
            )

    def _pandas_dtype_to_mdqm(str_type_raw: str) -> str:
        """Align with /jobs/*/upload column type labels."""
        dt = str(str_type_raw).lower()
        if "int" in dt:
            return "Integer"
        if "float" in dt:
            return "Float"
        if "datetime" in dt or "timestamp" in dt:
            return "Date"
        if "bool" in dt:
            return "Boolean"
        return "String"

    def _normalize_import_dataframe_dates(df):
        for col in df.columns:
            if df[col].dtype == "object":
                try:
                    converted = pd.to_datetime(
                        df[col], format="mixed", dayfirst=True, errors="coerce"
                    )
                    if not converted.isna().all():
                        df[col] = converted
                except Exception:
                    pass
        return df

    def _snapshot_dataframe_to_job_table(db: Session, job_id: int, table_id: int, table_name: str, df):
        """Write CSV snapshot and resync column metadata + clear stats (same idea as replace-from-path)."""
        final_csv_path = table_csv_path(job_id, table_name)
        df.to_csv(final_csv_path, index=False)
        tbl = (
            db.query(models.TableMetadata)
            .filter(
                models.TableMetadata.job_id == job_id,
                models.TableMetadata.table_id == table_id,
            )
            .first()
        )
        if tbl:
            tbl.row_count = int(len(df))
        db.query(models.ColumnMetadata).filter(
            models.ColumnMetadata.job_id == job_id,
            models.ColumnMetadata.table_id == table_id,
        ).delete(synchronize_session=False)
        for col_name, dtype in df.dtypes.items():
            str_type = _pandas_dtype_to_mdqm(str(dtype))
            db.add(
                models.ColumnMetadata(
                    job_id=job_id,
                    table_id=table_id,
                    column_name=str(col_name),
                    data_type=str_type,
                )
            )
        db.query(models.TableStats).filter(
            models.TableStats.job_id == job_id,
            models.TableStats.table_id == table_id,
        ).delete(synchronize_session=False)
        db.commit()

    def _resolve_import_creds_from_job(db: Session, job: models.Job, body: dict | None = None):
        """Build Postgres creds + schema/tables from job.db_source_config (saved connection aware)."""
        body = body or {}
        cfg = job.db_source_config
        if not cfg or not isinstance(cfg, dict) or cfg.get("kind") != "postgres_tables":
            raise HTTPException(
                status_code=400,
                detail="Import is only available for database-backed datasets.",
            )
        schema_name = str(cfg.get("schema_name") or "").strip()
        table_names = cfg.get("table_names") or []
        if not schema_name or not isinstance(table_names, list) or len(table_names) == 0:
            raise HTTPException(status_code=400, detail="Stored DB source configuration is incomplete.")

        cid_raw = cfg.get("connection_id")
        cid_norm = None
        if cid_raw is not None and str(cid_raw).strip() != "" and str(cid_raw).lower() not in ("null", "nan"):
            try:
                cid_norm = int(float(cid_raw))
            except (ValueError, TypeError):
                cid_norm = None
        has_saved = cid_norm is not None

        resolve_pl: dict = {
            "dbname": cfg.get("dbname"),
            "host": cfg.get("host"),
            "port": str(cfg.get("port") or "5432"),
            "user": cfg.get("user"),
            "db_type": cfg.get("db_type") or "postgres",
        }
        from utils.source_secret_crypto import decrypt_db_password_optional

        decrypted_stored = decrypt_db_password_optional(cfg.get("encrypted_db_pass"))
        if decrypted_stored:
            resolve_pl["pass"] = decrypted_stored
        if has_saved:
            resolve_pl["connection_id"] = cid_norm

        if str(body.get("host") or "").strip():
            resolve_pl["host"] = str(body["host"]).strip()
        if str(body.get("port") or "").strip():
            resolve_pl["port"] = str(body["port"]).strip()
        if str(body.get("user") or "").strip():
            resolve_pl["user"] = str(body["user"]).strip()
        if str(body.get("dbname") or "").strip():
            resolve_pl["dbname"] = str(body["dbname"]).strip()
        if body.get("pass") is not None and str(body.get("pass", "")) != "":
            resolve_pl["pass"] = body["pass"]

        if not has_saved and not str(resolve_pl.get("pass") or "").strip():
            raise HTTPException(
                status_code=400,
                detail="No stored password for this dataset. Re-save using a saved DB connection profile.",
            )

        try:
            creds = _resolve_connection_payload(resolve_pl)
        except HTTPException as exc:
            if exc.status_code != 404 or not has_saved:
                raise
            creds = {
                "host": cfg.get("host"),
                "port": str(cfg.get("port") or "5432"),
                "user": cfg.get("user"),
                "pass": "",
                "dbname": cfg.get("dbname"),
                "db_type": cfg.get("db_type") or "postgres",
            }
            if body.get("pass"):
                creds["pass"] = str(body["pass"])

        creds["host"] = str(creds.get("host") or cfg.get("host") or "").strip()
        creds["user"] = str(creds.get("user") or cfg.get("user") or "").strip()
        creds["dbname"] = str(creds.get("dbname") or cfg.get("dbname") or "").strip()
        creds["port"] = str(creds.get("port") or cfg.get("port") or "5432")
        creds["db_type"] = str(creds.get("db_type") or cfg.get("db_type") or "postgres").strip().lower()
        if not creds.get("host") or not creds.get("user") or not creds.get("dbname"):
            raise HTTPException(status_code=400, detail="Cannot resolve database credentials for import.")

        return creds, schema_name, [str(t).strip() for t in table_names if str(t).strip()]

    def _execute_import_for_job(job_id: int, body: dict | None = None):
        from services.dataset_db_import import import_tables_into_job

        db = SessionLocal()
        external_conn = None
        try:
            job = db.query(models.Job).filter(models.Job.job_id == job_id).first()
            if not job:
                return
            try:
                creds, schema_name, table_names = _resolve_import_creds_from_job(db, job, body)
            except HTTPException:
                job.status = "Import failed"
                db.commit()
                return
            job.status = "Importing"
            db.commit()
            external_conn = _connect_db(creds)
            import_tables_into_job(
                db,
                job_id=job_id,
                external_conn=external_conn,
                schema_name=schema_name,
                table_names=table_names,
                snapshot_fn=_snapshot_dataframe_to_job_table,
                db_type=creds.get("db_type") or "postgres",
            )
            job = db.query(models.Job).filter(models.Job.job_id == job_id).first()
            if job:
                job.status = "Pending"
                db.commit()
        except Exception:
            job = db.query(models.Job).filter(models.Job.job_id == job_id).first()
            if job:
                job.status = "Import failed"
                db.commit()
        finally:
            if external_conn is not None:
                try:
                    external_conn.close()
                except Exception:
                    pass
            db.close()

    @app.post("/db/register-dataset")
    def register_dataset_source(request: Request, payload: dict = Body(...), db: Session = Depends(get_db)):
        """
        Register a dataset job + single table without pulling rows yet.
        Data Owner uses /jobs/{id}/import-from-db to load data in the background.
        """
        from services.dataset_db_import import build_db_source_config

        job_name = str(payload.get("job_name") or "").strip()
        schema_name = str(payload.get("schema_name") or "").strip()
        raw_tables = payload.get("table_names")

        if not job_name:
            raise HTTPException(status_code=400, detail="job_name is required.")
        if not schema_name:
            raise HTTPException(status_code=400, detail="schema_name is required.")
        if not isinstance(raw_tables, list) or len(raw_tables) != 1:
            raise HTTPException(status_code=400, detail="Exactly one table must be selected.")

        table_name = str(raw_tables[0] or "").strip()
        if not table_name:
            raise HTTPException(status_code=400, detail="table_names[0] must be a non-empty string.")

        creds = _resolve_connection_payload(payload, getattr(request.state, "user_id", None))
        if not creds.get("host") or not creds.get("user") or not creds.get("dbname"):
            raise HTTPException(
                status_code=400,
                detail="host, user, and dbname are required (or provide connection_id + dbname).",
            )

        new_job = models.Job(job_name=job_name, status="Registered")
        db.add(new_job)
        db.commit()
        db.refresh(new_job)
        job_id = new_job.job_id

        new_table = models.TableMetadata(
            job_id=job_id,
            table_id=1,
            table_name=table_name,
            row_count=0,
        )
        db.add(new_table)
        db.commit()

        cfg_dict = build_db_source_config(payload, creds, schema_name, [table_name])
        jr = db.query(models.Job).filter(models.Job.job_id == job_id).first()
        if jr:
            jr.db_source_config = cfg_dict
            db.commit()

        return {
            "message": "Dataset registered (data not loaded yet). Run import when ready.",
            "job_id": job_id,
            "status": "Registered",
            "created_jobs": [{"job_id": job_id}],
        }

    @app.post("/jobs/{job_id}/import-from-db")
    def import_job_from_database(
        job_id: int,
        background_tasks: BackgroundTasks,
        body: Optional[dict] = Body(None),
        db: Session = Depends(get_db),
    ):
        job = db.query(models.Job).filter(models.Job.job_id == job_id).first()
        if not job:
            raise HTTPException(status_code=404, detail="Job not found")
        if job.status == "Importing":
            raise HTTPException(status_code=409, detail="Import already in progress for this job.")

        cfg = job.db_source_config
        if not cfg or not isinstance(cfg, dict) or cfg.get("kind") != "postgres_tables":
            raise HTTPException(
                status_code=400,
                detail="Import is only available for database-backed datasets.",
            )

        job.status = "Importing"
        db.commit()
        background_tasks.add_task(_execute_import_for_job, job_id, body or {})
        return {
            "message": "Import started in the background. You can close this screen.",
            "job_id": job_id,
            "status": "Importing",
        }

    @app.post("/db/connect")
    def connect_db_create_job(request: Request, payload: dict = Body(...), db: Session = Depends(get_db)):
        """
        Create one job and register one or more tables by snapshotting rows from an external Postgres DB into CSV uploads.
        Response shape matches the frontend (created_jobs[].job_id).
        """
        job_name = str(payload.get("job_name") or "").strip()
        schema_name = str(payload.get("schema_name") or "").strip()
        raw_tables = payload.get("table_names")

        if not job_name:
            raise HTTPException(status_code=400, detail="job_name is required.")
        if not schema_name:
            raise HTTPException(status_code=400, detail="schema_name is required.")
        if not isinstance(raw_tables, list) or len(raw_tables) == 0:
            raise HTTPException(status_code=400, detail="table_names must be a non-empty list.")

        table_names = []
        for t in raw_tables:
            s = str(t or "").strip()
            if not s:
                raise HTTPException(status_code=400, detail="Each table name must be a non-empty string.")
            table_names.append(s)

        seen = set()
        table_names_unique = []
        for t in table_names:
            if t not in seen:
                seen.add(t)
                table_names_unique.append(t)
        table_names = table_names_unique

        creds = _resolve_connection_payload(payload, getattr(request.state, "user_id", None), getattr(request.state, "user_id", None))
        if not creds.get("host") or not creds.get("user") or not creds.get("dbname"):
            raise HTTPException(
                status_code=400,
                detail="host, user, and dbname are required (or provide connection_id + dbname).",
            )

        # Require encryption when a DB password is in use so it can persist on the job for refresh-from-db.
        plain_pw = str(creds.get("pass") or "").strip()
        if plain_pw:
            from utils.source_secret_crypto import encrypt_db_password_optional, encryption_available

            if not encryption_available():
                raise HTTPException(
                    status_code=400,
                    detail=(
                        "Set MDQM_DB_SOURCE_MASTER_SECRET in the backend .env (see .env.example). "
                        "It is required so your database password can be encrypted and saved on the job for Refresh "
                        "without typing it again."
                    ),
                )
            enc_probe = encrypt_db_password_optional(plain_pw)
            if not enc_probe:
                raise HTTPException(
                    status_code=400,
                    detail=(
                        "Password encryption failed. Install the cryptography package in the backend "
                        "(e.g. pip install cryptography), then restart the API."
                    ),
                )

        external_conn = None

        try:
            new_job = models.Job(job_name=job_name, status="Pending")
            db.add(new_job)
            db.commit()
            db.refresh(new_job)
            job_id = new_job.job_id

            external_conn = _connect_db(creds)
            max_id = (
                db.query(func.max(models.TableMetadata.table_id))
                .filter(models.TableMetadata.job_id == job_id)
                .scalar()
            )
            next_table_id = 1 if max_id is None else int(max_id) + 1

            db_type = str(creds.get("db_type") or "postgres").lower().strip()
            for table_name in table_names:
                if db_type in ("mssql", "sqlserver", "sql_server"):
                    q_str = f"SELECT * FROM [{schema_name}].[{table_name}]"
                elif db_type == "mysql":
                    if schema_name:
                        q_str = f"SELECT * FROM `{schema_name}`.`{table_name}`"
                    else:
                        q_str = f"SELECT * FROM `{table_name}`"
                else:
                    query = psql.SQL("SELECT * FROM {}.{}").format(
                        psql.Identifier(schema_name),
                        psql.Identifier(table_name),
                    )
                    q_str = query.as_string(external_conn)
                df = pd.read_sql_query(q_str, external_conn)
                df = _normalize_import_dataframe_dates(df)

                new_table = models.TableMetadata(
                    job_id=job_id,
                    table_id=next_table_id,
                    table_name=table_name,
                    row_count=int(len(df)),
                )
                db.add(new_table)
                db.commit()

                _snapshot_dataframe_to_job_table(db, job_id, next_table_id, table_name, df)
                next_table_id += 1

            jr = db.query(models.Job).filter(models.Job.job_id == job_id).first()
            warnings = []
            if jr:
                from utils.source_secret_crypto import encrypt_db_password_optional, encryption_available

                cfg_dict = {
                    "kind": "postgres_tables",
                    "connection_id": payload.get("connection_id"),
                    "dbname": creds["dbname"],
                    "db_type": creds.get("db_type") or "postgres",
                    "schema_name": schema_name,
                    "table_names": list(table_names),
                    "host": creds.get("host"),
                    "port": str(creds.get("port") or "5432"),
                    "user": creds.get("user"),
                }
                enc = encrypt_db_password_optional(creds.get("pass") or "")
                if enc:
                    cfg_dict["encrypted_db_pass"] = enc
                jr.db_source_config = cfg_dict
                try:
                    db.commit()
                except Exception as cfg_exc:
                    db.rollback()
                    warnings.append(
                        "Job and tables were created, but storing refresh settings failed. "
                        f"Restart the API (and ensure column metadata.jobs.db_source_config exists). Detail: {cfg_exc}"
                    )

            resp = {
                "message": "Connected and imported tables",
                "created_jobs": [{"job_id": job_id}],
            }
            if warnings:
                resp["warnings"] = warnings
            return resp
        except HTTPException:
            raise
        except Exception as e:
            raise HTTPException(status_code=400, detail=f"Database import failed: {str(e)}")
        finally:
            if external_conn is not None:
                try:
                    external_conn.close()
                except Exception:
                    pass

    @app.post("/jobs/{job_id}/refresh-from-db")
    def refresh_job_from_database(
        job_id: int,
        body: Optional[dict] = Body(None),
        db: Session = Depends(get_db),
    ):
        """
        Re-query the external database using the connection info stored on the job and
        rewrite uploads/*.csv plus column metadata for each registered table.
        """
        body = body or {}
        job = db.query(models.Job).filter(models.Job.job_id == job_id).first()
        if not job:
            raise HTTPException(status_code=404, detail="Job not found")
        cfg = job.db_source_config
        if not cfg or not isinstance(cfg, dict) or cfg.get("kind") != "postgres_tables":
            raise HTTPException(
                status_code=400,
                detail="Refresh is only available for database-backed datasets (Data Owner → Table).",
            )
        schema_name = str(cfg.get("schema_name") or "").strip()
        table_names = cfg.get("table_names") or []
        if not schema_name or not isinstance(table_names, list) or len(table_names) == 0:
            raise HTTPException(status_code=400, detail="Stored DB source configuration is incomplete.")

        # connection_id from JSON may be int, float, or string
        cid_raw = cfg.get("connection_id")
        cid_norm = None
        if cid_raw is not None and str(cid_raw).strip() != "" and str(cid_raw).lower() not in ("null", "nan"):
            try:
                cid_norm = int(float(cid_raw))
            except (ValueError, TypeError):
                cid_norm = None
        has_saved = cid_norm is not None

        db_type = str(cfg.get("db_type") or "postgres").lower().strip()
        default_port = "1433" if db_type in ("mssql", "sqlserver", "sql_server") else ("3306" if db_type == "mysql" else "5432")

        resolve_pl: dict = {
            "dbname": cfg.get("dbname"),
            "host": cfg.get("host"),
            "port": str(cfg.get("port") or default_port),
            "user": cfg.get("user"),
            "db_type": db_type,
        }

        from utils.source_secret_crypto import decrypt_db_password_optional

        decrypted_stored = decrypt_db_password_optional(cfg.get("encrypted_db_pass"))
        if decrypted_stored:
            resolve_pl["pass"] = decrypted_stored

        if has_saved:
            resolve_pl["connection_id"] = cid_norm
        else:
            resolve_pl.pop("connection_id", None)

        if str(body.get("host") or "").strip():
            resolve_pl["host"] = str(body["host"]).strip()
        if str(body.get("port") or "").strip():
            resolve_pl["port"] = str(body["port"]).strip()
        if str(body.get("user") or "").strip():
            resolve_pl["user"] = str(body["user"]).strip()
        if str(body.get("dbname") or "").strip():
            resolve_pl["dbname"] = str(body["dbname"]).strip()
        if body.get("pass") is not None and str(body.get("pass", "")) != "":
            resolve_pl["pass"] = body["pass"]

        if not has_saved:
            if not str(resolve_pl.get("pass") or "").strip():
                raise HTTPException(
                    status_code=400,
                    detail="Database password missing: enable MDQM_DB_SOURCE_MASTER_SECRET and re-import once to store "
                    'it encrypted, or pass {"pass": "..."} in the refresh request body.',
                )

        try:
            creds = _resolve_connection_payload(resolve_pl)
        except HTTPException as exc:
            if exc.status_code != 404 or not has_saved:
                raise
            creds = {
                "host": cfg.get("host"),
                "port": str(cfg.get("port") or default_port),
                "user": cfg.get("user"),
                "pass": "",
                "dbname": cfg.get("dbname"),
                "db_type": db_type,
            }
            pb = body.get("pass")
            if pb is not None and str(pb) != "":
                creds["pass"] = str(pb)

        def _pick(*vals):
            for v in vals:
                s = str(v or "").strip()
                if s:
                    return s
            return ""

        creds["host"] = _pick(creds.get("host"), cfg.get("host"))
        creds["user"] = _pick(creds.get("user"), cfg.get("user"))
        creds["dbname"] = _pick(creds.get("dbname"), cfg.get("dbname"))
        creds["port"] = str(creds.get("port") or cfg.get("port") or default_port)
        creds["db_type"] = _pick(creds.get("db_type"), cfg.get("db_type"), db_type)
        pb = body.get("pass")
        if pb is not None and str(pb) != "":
            creds["pass"] = str(pb)

        if not creds.get("host") or not creds.get("user") or not creds.get("dbname"):
            raise HTTPException(
                status_code=400,
                detail="Cannot resolve database credentials for refresh. Enter the database password in the box below "
                "(required if the saved connection has no password on disk, or if the saved connection was removed). "
                "You can also re-create the dataset from Data Owner → Table.",
            )

        tms = db.query(models.TableMetadata).filter(models.TableMetadata.job_id == job_id).all()
        name_to_tm = {t.table_name: t for t in tms}

        external_conn = None
        summaries = []
        try:
            from services.dataset_db_import import import_tables_into_job

            # Verify that all tables in configuration are registered
            for table_name in table_names:
                tn = str(table_name).strip()
                tm = name_to_tm.get(tn)
                if not tm:
                    raise HTTPException(
                        status_code=400,
                        detail=f"Table '{tn}' is not registered on this job; cannot refresh.",
                    )

            external_conn = _connect_db(creds)
            summaries = import_tables_into_job(
                db,
                job_id=job_id,
                external_conn=external_conn,
                schema_name=schema_name,
                table_names=table_names,
                snapshot_fn=_snapshot_dataframe_to_job_table,
                db_type=creds.get("db_type") or "postgres",
            )
        except HTTPException:
            raise
        except Exception as e:
            raise HTTPException(status_code=400, detail=f"Refresh failed: {str(e)}")
        finally:
            if external_conn is not None:
                try:
                    external_conn.close()
                except Exception:
                    pass

        return {
            "message": "Job tables refreshed from database",
            "job_id": job_id,
            "tables": summaries,
        }

    @app.post("/db/list-schemas-tables")
    def list_schemas_tables(request: Request, payload: dict = Body(...)):
        creds = _resolve_connection_payload(payload, getattr(request.state, "user_id", None))
        if not creds.get("host") or not creds.get("user") or not creds.get("dbname"):
            raise HTTPException(
                status_code=400,
                detail="Provide database name, host, and username (or fix the saved connection), then try again.",
            )

        conn = None
        try:
            conn = _connect_db(creds)
            cur = conn.cursor()
            db_type = str(creds.get("db_type") or "postgres").lower().strip()

            if db_type in ("mssql", "sqlserver", "sql_server"):
                cur.execute(
                    """
                    SELECT name 
                    FROM sys.schemas 
                    WHERE name NOT IN ('sys', 'db_owner', 'db_accessadmin', 'db_securityadmin', 'db_ddladmin', 'db_backupoperator', 'db_datareader', 'db_datawriter', 'db_denydatareader', 'db_denydatawriter', 'guest', 'INFORMATION_SCHEMA') 
                    ORDER BY name
                    """
                )
                schemas = [row[0] for row in cur.fetchall()]

                tables_by_schema = {}
                for schema in schemas:
                    cur.execute(
                        """
                        SELECT name 
                        FROM sys.tables 
                        WHERE schema_id = SCHEMA_ID(?)
                        ORDER BY name
                        """,
                        (schema,),
                    )
                    tables_by_schema[schema] = [row[0] for row in cur.fetchall()]
            elif db_type == "mysql":
                dbname = creds.get("dbname")
                schemas = [dbname]
                tables_by_schema = {}
                cur.execute(
                    """
                    SELECT table_name 
                    FROM information_schema.tables 
                    WHERE table_schema = %s 
                    AND table_type = 'BASE TABLE'
                    ORDER BY table_name
                    """,
                    (dbname,),
                )
                tables_by_schema[dbname] = [row[0] for row in cur.fetchall()]
            elif db_type == "oracle":
                cur.execute(
                    """
                    SELECT username 
                    FROM all_users 
                    WHERE oracle_maintained = 'N' 
                    ORDER BY username
                    """
                )
                schemas = [row[0] for row in cur.fetchall()]
                user_schema = str(creds["user"]).upper()
                if user_schema not in schemas:
                    schemas.insert(0, user_schema)
                
                tables_by_schema = {}
                for schema in schemas:
                    cur.execute(
                        """
                        SELECT table_name 
                        FROM all_tables 
                        WHERE owner = :1 
                        ORDER BY table_name
                        """,
                        (schema,),
                    )
                    tables_by_schema[schema] = [row[0] for row in cur.fetchall()]
            elif db_type == "snowflake":
                cur.execute(
                    """
                    SELECT schema_name
                    FROM information_schema.schemata
                    WHERE schema_name NOT IN ('INFORMATION_SCHEMA')
                    ORDER BY schema_name
                    """
                )
                schemas = [row[0] for row in cur.fetchall()]

                tables_by_schema = {}
                for schema in schemas:
                    cur.execute(
                        """
                        SELECT table_name
                        FROM information_schema.tables
                        WHERE table_schema = %s
                        AND table_type = 'BASE TABLE'
                        ORDER BY table_name
                        """,
                        (schema,),
                    )
                    tables_by_schema[schema] = [row[0] for row in cur.fetchall()]
            elif db_type == "databricks":
                cur.execute("SHOW SCHEMAS")
                schemas = [row[0] for row in cur.fetchall() if row[0] != "information_schema"]
                tables_by_schema = {}
                for schema in schemas:
                    cur.execute(f"SHOW TABLES IN `{schema}`")
                    tables_by_schema[schema] = [row[1] for row in cur.fetchall()]
            else:
                cur.execute(
                    """
                    SELECT schema_name
                    FROM information_schema.schemata
                    WHERE schema_name NOT IN ('pg_catalog', 'information_schema')
                    ORDER BY schema_name
                    """
                )
                schemas = [row[0] for row in cur.fetchall()]

                tables_by_schema = {}
                for schema in schemas:
                    cur.execute(
                        """
                        SELECT table_name
                        FROM information_schema.tables
                        WHERE table_schema = %s
                        AND table_type = 'BASE TABLE'
                        ORDER BY table_name
                        """,
                        (schema,),
                    )
                    tables_by_schema[schema] = [row[0] for row in cur.fetchall()]

            cur.close()
            return {"schemas": schemas, "tables_by_schema": tables_by_schema}
        except Exception as e:
            raise HTTPException(status_code=400, detail=f"Failed to list schemas/tables: {str(e)}")
        finally:
            if conn is not None:
                conn.close()


    @app.post("/db/table-columns")
    def db_table_columns(request: Request, payload: dict = Body(...)):
        creds = _resolve_connection_payload(payload, getattr(request.state, "user_id", None))
        schema_name = payload.get("schema_name")
        table_name = payload.get("table_name")

        if not creds.get("host") or not creds.get("user") or not creds.get("dbname"):
            raise HTTPException(
                status_code=400,
                detail="host, user, and dbname are required (or provide connection_id + dbname)",
            )
        if not schema_name or not table_name:
            raise HTTPException(status_code=400, detail="schema_name and table_name are required")

        conn = None
        try:
            conn = _connect_db(creds)
            cur = conn.cursor()
            db_type = str(creds.get("db_type") or "postgres").lower().strip()

            if db_type in ("mssql", "sqlserver", "sql_server"):
                cur.execute(
                    """
                    SELECT column_name
                    FROM information_schema.columns
                    WHERE table_schema = ?
                    AND table_name = ?
                    ORDER BY ordinal_position
                    """,
                    (schema_name, table_name),
                )
            elif db_type == "oracle":
                cur.execute(
                    """
                    SELECT column_name
                    FROM all_tab_columns
                    WHERE owner = :1
                    AND table_name = :2
                    ORDER BY column_id
                    """,
                    (schema_name.upper(), table_name.upper()),
                )
            elif db_type == "databricks":
                cur.execute(f"DESCRIBE TABLE `{schema_name}`.`{table_name}`")
            else:
                cur.execute(
                    """
                    SELECT column_name
                    FROM information_schema.columns
                    WHERE table_schema = %s
                    AND table_name = %s
                    ORDER BY ordinal_position
                    """,
                    (schema_name, table_name),
                )
            if db_type == "databricks":
                columns = [row[0] for row in cur.fetchall() if row[0] and not row[0].startswith("#") and row[0].strip()]
            else:
                columns = [row[0] for row in cur.fetchall()]
            cur.close()
            return {"columns": columns}
        except Exception as e:
            raise HTTPException(status_code=400, detail=f"Failed to fetch table columns: {str(e)}")
        finally:
            if conn is not None:
                conn.close()


    @app.post("/db/lookup-values")
    def db_lookup_values(request: Request, payload: dict = Body(...)):
        creds = _resolve_connection_payload(payload, getattr(request.state, "user_id", None))
        schema_name = payload.get("schema_name")
        table_name = payload.get("table_name")
        column_name = payload.get("column_name")

        if not creds.get("host") or not creds.get("user") or not creds.get("dbname"):
            raise HTTPException(
                status_code=400,
                detail="host, user, and dbname are required (or provide connection_id + dbname)",
            )
        if not schema_name or not table_name or not column_name:
            raise HTTPException(
                status_code=400,
                detail="schema_name, table_name, and column_name are required",
            )

        conn = None
        try:
            conn = _connect_db(creds)
            cur = conn.cursor()
            db_type = str(creds.get("db_type") or "postgres").lower().strip()

            if db_type in ("mssql", "sqlserver", "sql_server"):
                c = column_name.replace(']', ']]')
                s = schema_name.replace(']', ']]')
                t = table_name.replace(']', ']]')
                query = f"""
                    SELECT DISTINCT CAST([{c}] AS NVARCHAR(MAX)) AS value
                    FROM [{s}].[{t}]
                    WHERE [{c}] IS NOT NULL
                    ORDER BY value
                """
                cur.execute(query)
            elif db_type == "mysql":
                c = column_name.replace('`', '``')
                s = schema_name.replace('`', '``')
                t = table_name.replace('`', '``')
                if s:
                    query = f"""
                        SELECT DISTINCT CAST(`{c}` AS CHAR) AS value
                        FROM `{s}`.`{t}`
                        WHERE `{c}` IS NOT NULL
                        ORDER BY value
                    """
                else:
                    query = f"""
                        SELECT DISTINCT CAST(`{c}` AS CHAR) AS value
                        FROM `{t}`
                        WHERE `{c}` IS NOT NULL
                        ORDER BY value
                    """
                cur.execute(query)
            elif db_type == "oracle":
                c = column_name.replace('"', '""')
                s = schema_name.replace('"', '""')
                t = table_name.replace('"', '""')
                if s:
                    query = f'SELECT DISTINCT CAST("{c}" AS VARCHAR2(4000)) AS value FROM "{s}"."{t}" WHERE "{c}" IS NOT NULL ORDER BY value'
                else:
                    query = f'SELECT DISTINCT CAST("{c}" AS VARCHAR2(4000)) AS value FROM "{t}" WHERE "{c}" IS NOT NULL ORDER BY value'
                cur.execute(query)
            elif db_type == "snowflake":
                c = column_name.replace('"', '""')
                s = schema_name.replace('"', '""')
                t = table_name.replace('"', '""')
                if s:
                    query = f'SELECT DISTINCT CAST("{c}" AS VARCHAR) AS value FROM "{s}"."{t}" WHERE "{c}" IS NOT NULL ORDER BY value'
                else:
                    query = f'SELECT DISTINCT CAST("{c}" AS VARCHAR) AS value FROM "{t}" WHERE "{c}" IS NOT NULL ORDER BY value'
                cur.execute(query)
            elif db_type == "databricks":
                c = column_name.replace('`', '``')
                s = schema_name.replace('`', '``')
                t = table_name.replace('`', '``')
                if s:
                    query = f'SELECT DISTINCT CAST(`{c}` AS STRING) AS value FROM `{s}`.`{t}` WHERE `{c}` IS NOT NULL ORDER BY value'
                else:
                    query = f'SELECT DISTINCT CAST(`{c}` AS STRING) AS value FROM `{t}` WHERE `{c}` IS NOT NULL ORDER BY value'
                cur.execute(query)
            else:
                query = psql.SQL(
                    """
                    SELECT DISTINCT CAST({col} AS TEXT) AS value
                    FROM {schema}.{table}
                    WHERE {col} IS NOT NULL
                    ORDER BY value
                    """
                ).format(
                    col=psql.Identifier(column_name),
                    schema=psql.Identifier(schema_name),
                    table=psql.Identifier(table_name),
                )
                cur.execute(query)
            values = [str(row[0]).strip() for row in cur.fetchall() if str(row[0]).strip()]
            cur.close()
            return {"values": values}
        except Exception as e:
            raise HTTPException(status_code=400, detail=f"Failed to fetch lookup values: {str(e)}")
        finally:
            if conn is not None:
                conn.close()


    @app.post("/db/export-results")
    def export_results_to_external_db(request: Request, payload: dict = Body(...), db: Session = Depends(get_db)):
        """
        Export combined result rows (clean + error) into an external Postgres table.
        Supports either connection_id + dbname OR direct host/port/user/pass/dbname.
        """
        table_id = payload.get("table_id")
        job_id = payload.get("job_id")
        target_table = payload.get("target_table")
        target_schema = payload.get("target_schema") or None
        if_exists = payload.get("if_exists", "append")

        if not table_id or not target_table:
            raise HTTPException(status_code=400, detail="table_id and target_table are required")
        if if_exists not in ("append", "replace"):
            raise HTTPException(status_code=400, detail="if_exists must be 'append' or 'replace'")

        # Resolve source table safely. table_id is not globally unique across jobs.
        if job_id is not None:
            src_table = (
                db.query(models.TableMetadata)
                .filter(
                    models.TableMetadata.job_id == int(job_id),
                    models.TableMetadata.table_id == table_id,
                )
                .first()
            )
        else:
            # Fallback for older clients: choose the newest table_id that actually has stats.
            src_table = (
                db.query(models.TableMetadata)
                .join(
                    models.TableStats,
                    (models.TableStats.job_id == models.TableMetadata.job_id)
                    & (models.TableStats.table_id == models.TableMetadata.table_id),
                )
                .filter(models.TableMetadata.table_id == table_id)
                .order_by(models.TableStats.stat_id.desc())
                .first()
            )
            if not src_table:
                src_table = (
                    db.query(models.TableMetadata)
                    .filter(models.TableMetadata.table_id == table_id)
                    .order_by(models.TableMetadata.job_id.desc())
                    .first()
                )
        if not src_table:
            raise HTTPException(status_code=404, detail="Source table not found")

        creds = _resolve_connection_payload(payload, getattr(request.state, "user_id", None), getattr(request.state, "user_id", None))
        if not creds.get("host") or not creds.get("user") or not creds.get("dbname"):
            raise HTTPException(
                status_code=400,
                detail="host, user, and dbname are required (or provide connection_id + dbname)",
            )
        # Enforce split workflow: internal MDQM metadata stays in POSTGRES_DB (usually mdms),
        # while exported result tables must go to a separate target database.
        internal_dbname = str(POSTGRES_DB).strip().lower()
        target_dbname = str(creds.get("dbname", "")).strip().lower()
        allow_internal_export = bool(payload.get("allow_internal_export_db", False))
        if target_dbname == internal_dbname and not allow_internal_export:
            raise HTTPException(
                status_code=400,
                detail=(
                    f"Target database '{creds.get('dbname')}' is reserved for MDQM backend metadata. "
                    "Choose a separate output database (for example 'postgres') for exported CSV results."
                ),
            )

        # Keep the same resilience pattern used by other DB helper endpoints.
        conn = None
        try:
            conn = _connect_db(creds)
            conn.close()
        except Exception as e:
            if conn is not None:
                conn.close()
            raise HTTPException(status_code=400, detail=f"Failed to connect target DB: {str(e)}")

        try:
            try:
                df = generate_output_dataframe(db, src_table.table_name, src_table.job_id)
            except Exception as e:
                # Fallback: some runs may not materialize app_data clean/error tables
                # (for example, identifier truncation on very long table names).
                # In that case export the latest uploaded source CSV instead of failing hard.
                source_csv_path = resolve_table_csv_path(src_table.job_id, src_table.table_name)
                if not source_csv_path:
                    raise e
                df = pd.read_csv(source_csv_path)

            db_type = str(creds.get("db_type") or "postgres").lower().strip()
            if db_type in ("mssql", "sqlserver", "sql_server"):
                ext_engine = create_engine("mssql+pyodbc://", creator=lambda: _connect_db(creds))
            elif db_type == "mysql":
                ext_engine = create_engine("mysql+pymysql://", creator=lambda: _connect_db(creds))
            else:
                ext_engine = create_engine("postgresql+psycopg2://", creator=lambda: _connect_db(creds))

            # Client-friendly guard for append mode:
            # If target table exists but schema differs, return a clear message
            # instead of a large raw SQL failure.
            if if_exists == "append":
                inspector = inspect(ext_engine)
                if inspector.has_table(target_table, schema=target_schema):
                    existing_cols = [
                        str(c.get("name"))
                        for c in inspector.get_columns(target_table, schema=target_schema)
                    ]
                    incoming_cols = [str(c) for c in df.columns.tolist()]
                    existing_set = set(existing_cols)
                    incoming_set = set(incoming_cols)
                    missing_in_target = sorted(incoming_set - existing_set)
                    extra_in_target = sorted(existing_set - incoming_set)
                    if missing_in_target or extra_in_target:
                        msg_parts = [
                            "Schema mismatch for append mode.",
                            f"Target table '{target_table}' has a different structure.",
                        ]
                        if missing_in_target:
                            msg_parts.append(
                                f"Missing in target: {', '.join(missing_in_target[:12])}"
                                + (" ..." if len(missing_in_target) > 12 else "")
                            )
                        if extra_in_target:
                            msg_parts.append(
                                f"Extra in target: {', '.join(extra_in_target[:12])}"
                                + (" ..." if len(extra_in_target) > 12 else "")
                            )
                        msg_parts.append(
                            "Use 'replace' for full reload, or export to a new target table for this dataset."
                        )
                        raise HTTPException(status_code=400, detail=" ".join(msg_parts))

            df.to_sql(
                name=target_table,
                con=ext_engine,
                schema=target_schema,
                if_exists=if_exists,
                index=False,
                method="multi",
                chunksize=2000,
            )
            return {
                "message": "Results exported successfully",
                "rows_exported": int(len(df)),
                "target_dbname": creds["dbname"],
                "target_host": creds["host"],
                "target_schema": target_schema,
                "target_table": target_table,
            }
        except Exception as e:
            raise HTTPException(status_code=400, detail=f"Failed to export results: {str(e)}")

    @app.post("/jobs/create")
    def create_job(payload: JobCreate, db: Session = Depends(get_db)):
        # Notice we now use payload.job_name
        new_job = models.Job(job_name=payload.job_name, status="Pending")
        db.add(new_job)
        db.commit()
        db.refresh(new_job)
        return {"job_id": new_job.job_id, "message": "Job Created"}

    # 1. FIX: Changed the URL to match the frontend exactly
    @app.post("/jobs/{job_id}/upload")
    async def upload_file(
        job_id: int,
        file: UploadFile = File(...),
        source_path: Optional[str] = Form(None),
        db: Session = Depends(get_db),
    ):
        # Clean the table name upfront by removing any extension
        table_name = file.filename.replace(".csv", "").replace(".xlsx", "").replace(".xls", "")
        
        # Per-job CSV so the same filename on another job does not overwrite this one
        final_csv_path = table_csv_path(job_id, table_name)
        temp_file_path = job_temp_upload_path(job_id, file.filename)
        
        # Save the uploaded file temporarily
        with open(temp_file_path, "wb") as buffer:
            shutil.copyfileobj(file.file, buffer)
        
        max_id = db.query(func.max(models.TableMetadata.table_id)).filter(models.TableMetadata.job_id == job_id).scalar()
        next_table_id = 1 if max_id is None else max_id + 1

        try:
            # 2. FIX: Handle both Excel and CSV formats
            if file.filename.endswith(".xlsx") or file.filename.endswith(".xls"):
                # Read the Excel file and instantly save it as our standardized CSV
                df = pd.read_excel(temp_file_path)
                df.to_csv(final_csv_path, index=False)
                
                # Delete the original Excel file to save space
                if os.path.exists(temp_file_path):
                    os.remove(temp_file_path)
            else:
                df = pd.read_csv(final_csv_path)
                
            # ==========================================================
            # 3. FIX: THE MIXED-FORMAT DATE DETECTION MAGIC GOES HERE
            # ==========================================================
            for col in df.columns:
                if df[col].dtype == 'object':  # If Pandas thinks it's a generic String
                    try:
                        # format='mixed' handles changing formats in the same column!
                        converted = pd.to_datetime(df[col], format='mixed', dayfirst=True, errors='coerce')
                        
                        # If it successfully found real dates (not just empty/NaT), apply it!
                        if not converted.isna().all():
                            df[col] = converted
                    except Exception:
                        pass
            # ==========================================================

            row_count = len(df)
            columns = df.dtypes.items()
        except Exception as e:
            raise HTTPException(status_code=400, detail=f"Invalid File format: {str(e)}")

        new_table = models.TableMetadata(
            job_id=job_id,
            table_id=next_table_id,
            table_name=table_name,
            row_count=row_count 
        )
        db.add(new_table)
        db.commit()

        for col_name, dtype in columns:
            str_type = "String"
            if "int" in str(dtype): str_type = "Integer"
            elif "float" in str(dtype): str_type = "Float"
            elif "datetime" in str(dtype): str_type = "Date"
            elif "bool" in str(dtype): str_type = "Boolean"

            col_meta = models.ColumnMetadata(
                job_id=job_id,
                table_id=next_table_id,
                column_name=col_name,
                data_type=str_type
            )
            db.add(col_meta)
        db.commit()

        normalized_source_path = _normalize_local_path(source_path or "")
        if normalized_source_path and os.path.isfile(normalized_source_path):
            _save_table_source_path(job_id, next_table_id, normalized_source_path)

        return {"job_id": job_id, "message": "File Uploaded and Processed Successfully"}


    @app.post("/jobs/{job_id}/upload-from-path")
    def upload_file_from_path(job_id: int, payload: dict = Body(...), db: Session = Depends(get_db)):
        file_path = _normalize_local_path(payload.get("file_path"))
        if not file_path:
            raise HTTPException(status_code=400, detail="file_path is required")
        if not os.path.isfile(file_path):
            raise HTTPException(
                status_code=400,
                detail=f"File path not found. Use a full path like C:\\Downloads\\data.csv. Received: {file_path}",
            )

        file_name = os.path.basename(file_path)
        table_name = file_name.replace(".csv", "").replace(".xlsx", "").replace(".xls", "")
        final_csv_path = table_csv_path(job_id, table_name)

        max_id = db.query(func.max(models.TableMetadata.table_id)).filter(models.TableMetadata.job_id == job_id).scalar()
        next_table_id = 1 if max_id is None else max_id + 1

        try:
            if file_name.lower().endswith((".xlsx", ".xls")):
                df = pd.read_excel(file_path)
            else:
                df = pd.read_csv(file_path)

            # Normalize to job-scoped CSV for downstream engine.
            df.to_csv(final_csv_path, index=False)

            for col in df.columns:
                if df[col].dtype == "object":
                    try:
                        converted = pd.to_datetime(df[col], format="mixed", dayfirst=True, errors="coerce")
                        if not converted.isna().all():
                            df[col] = converted
                    except Exception:
                        pass

            row_count = len(df)
            columns = df.dtypes.items()
        except Exception as e:
            raise HTTPException(status_code=400, detail=f"Invalid file format: {str(e)}")

        new_table = models.TableMetadata(
            job_id=job_id,
            table_id=next_table_id,
            table_name=table_name,
            row_count=row_count,
        )
        db.add(new_table)
        db.commit()

        for col_name, dtype in columns:
            str_type = "String"
            if "int" in str(dtype):
                str_type = "Integer"
            elif "float" in str(dtype):
                str_type = "Float"
            elif "datetime" in str(dtype):
                str_type = "Date"
            elif "bool" in str(dtype):
                str_type = "Boolean"

            col_meta = models.ColumnMetadata(
                job_id=job_id,
                table_id=next_table_id,
                column_name=col_name,
                data_type=str_type,
            )
            db.add(col_meta)
        db.commit()
        _save_table_source_path(job_id, next_table_id, file_path)

        return {"job_id": job_id, "message": "File Uploaded from path and Processed Successfully"}

    @app.post("/jobs/{job_id}/tables/{table_id}/replace-from-path")
    def replace_table_file_from_path(job_id: int, table_id: int, payload: dict = Body(...), db: Session = Depends(get_db)):
        file_path = _normalize_local_path(payload.get("file_path"))
        if not file_path:
            raise HTTPException(status_code=400, detail="file_path is required")
        if not os.path.isfile(file_path):
            raise HTTPException(status_code=400, detail=f"File path not found: {file_path}")

        table = db.query(models.TableMetadata).filter(
            models.TableMetadata.job_id == job_id,
            models.TableMetadata.table_id == table_id,
        ).first()
        if not table:
            raise HTTPException(status_code=404, detail="Table not found for this job")

        final_csv_path = table_csv_path(job_id, table.table_name)

        try:
            if file_path.lower().endswith((".xlsx", ".xls")):
                df = pd.read_excel(file_path)
            else:
                df = pd.read_csv(file_path)
            df.to_csv(final_csv_path, index=False)
        except Exception as e:
            raise HTTPException(status_code=400, detail=f"Invalid file format: {str(e)}")

        # Refresh table metadata and columns to match replaced source.
        table.row_count = int(len(df))
        db.query(models.ColumnMetadata).filter(
            models.ColumnMetadata.job_id == job_id,
            models.ColumnMetadata.table_id == table_id,
        ).delete(synchronize_session=False)

        for col_name, dtype in df.dtypes.items():
            str_type = "String"
            if "int" in str(dtype):
                str_type = "Integer"
            elif "float" in str(dtype):
                str_type = "Float"
            elif "datetime" in str(dtype):
                str_type = "Date"
            elif "bool" in str(dtype):
                str_type = "Boolean"
            db.add(models.ColumnMetadata(
                job_id=job_id,
                table_id=table_id,
                column_name=col_name,
                data_type=str_type,
            ))

        # Clear old computed stats so UI doesn't keep old totals before next run.
        db.query(models.TableStats).filter(
            models.TableStats.job_id == job_id,
            models.TableStats.table_id == table_id,
        ).delete(synchronize_session=False)

        db.commit()
        _save_table_source_path(job_id, table_id, file_path)
        return {
            "message": "Table source file replaced successfully",
            "job_id": job_id,
            "table_id": table_id,
            "row_count": int(len(df)),
        }

    @app.post("/jobs/{job_id}/tables/{table_id}/replace-file")
    async def replace_table_file_upload(
        job_id: int,
        table_id: int,
        file: UploadFile = File(...),
        source_path: Optional[str] = Form(None),
        db: Session = Depends(get_db),
    ):
        table = db.query(models.TableMetadata).filter(
            models.TableMetadata.job_id == job_id,
            models.TableMetadata.table_id == table_id,
        ).first()
        if not table:
            raise HTTPException(status_code=404, detail="Table not found for this job")

        temp_file_path = job_temp_upload_path(job_id, file.filename)
        final_csv_path = table_csv_path(job_id, table.table_name)

        with open(temp_file_path, "wb") as buffer:
            shutil.copyfileobj(file.file, buffer)

        try:
            if file.filename.lower().endswith((".xlsx", ".xls")):
                df = pd.read_excel(temp_file_path)
            else:
                df = pd.read_csv(temp_file_path)
            df.to_csv(final_csv_path, index=False)
        except Exception as e:
            raise HTTPException(status_code=400, detail=f"Invalid file format: {str(e)}")
        finally:
            if os.path.exists(temp_file_path):
                os.remove(temp_file_path)

        table.row_count = int(len(df))
        db.query(models.ColumnMetadata).filter(
            models.ColumnMetadata.job_id == job_id,
            models.ColumnMetadata.table_id == table_id,
        ).delete(synchronize_session=False)

        for col_name, dtype in df.dtypes.items():
            str_type = "String"
            if "int" in str(dtype):
                str_type = "Integer"
            elif "float" in str(dtype):
                str_type = "Float"
            elif "datetime" in str(dtype):
                str_type = "Date"
            elif "bool" in str(dtype):
                str_type = "Boolean"
            db.add(models.ColumnMetadata(
                job_id=job_id,
                table_id=table_id,
                column_name=col_name,
                data_type=str_type,
            ))

        db.query(models.TableStats).filter(
            models.TableStats.job_id == job_id,
            models.TableStats.table_id == table_id,
        ).delete(synchronize_session=False)

        normalized_source_path = _normalize_local_path(source_path or "")
        if normalized_source_path and os.path.isfile(normalized_source_path):
            _save_table_source_path(job_id, table_id, normalized_source_path)

        db.commit()
        return {
            "message": "Table source file replaced successfully",
            "job_id": job_id,
            "table_id": table_id,
            "row_count": int(len(df)),
        }

    @app.post("/jobs/{job_id}/run")
    def run_job(job_id: int, db: Session = Depends(get_db)):
        # 1. Check if the job actually exists
        job = db.query(models.Job).filter(models.Job.job_id == job_id).first()
        if not job:
            raise HTTPException(status_code=404, detail="Job not found")
        
        # 2. Trigger the Python Engine!
        try:
            tables = db.query(models.TableMetadata).filter(models.TableMetadata.job_id == job_id).all()
            for t in tables:
                _refresh_table_csv_from_source_path(job_id, t.table_id, db)
            run_data_quality_job(job_id, db)
            return {"message": f"Job {job_id} executed successfully!"}
        except Exception as e:
            raise HTTPException(status_code=500, detail=f"Engine failed: {str(e)}")


    @app.post("/schedule-job/{job_id}")
    def schedule_job(job_id: int, data: dict = Body(...), db: Session = Depends(get_db)):
        job = db.query(models.Job).filter(models.Job.job_id == job_id).first()
        if not job:
            raise HTTPException(status_code=404, detail="Job not found")

        schedule_type = str(data.get("type", "")).strip().lower()
        if not schedule_type:
            raise HTTPException(status_code=400, detail="type is required")

        action = str(data.get("action", "dq")).strip().lower()
        run_fn = _run_scheduled_refresh if action == "refresh" else _run_scheduled_job

        job_key = _scheduler_job_key(job_id)
        if action == "refresh":
            job_key = f"{job_key}_refresh"
        try:
            scheduler.remove_job(job_key)
        except Exception:
            pass

        try:
            if schedule_type == "daily":
                hour, minute = map(int, str(data.get("time", "")).split(":"))
                scheduler.add_job(
                    run_fn,
                    "cron",
                    id=job_key,
                    replace_existing=True,
                    hour=hour,
                    minute=minute,
                    args=[job_id],
                )

            elif schedule_type == "weekly":
                hour, minute = map(int, str(data.get("time", "")).split(":"))
                scheduler.add_job(
                    run_fn,
                    "cron",
                    id=job_key,
                    replace_existing=True,
                    day_of_week=str(data.get("day", "0")),
                    hour=hour,
                    minute=minute,
                    args=[job_id],
                )

            elif schedule_type == "hourly":
                interval = max(int(data.get("interval", 1)), 1)
                scheduler.add_job(
                    run_fn,
                    "interval",
                    id=job_key,
                    replace_existing=True,
                    hours=interval,
                    args=[job_id],
                )

            elif schedule_type == "once":
                date_value = str(data.get("date", "")).strip()
                time_value = str(data.get("time", "")).strip()
                run_date = date_value
                if date_value and time_value and "T" not in date_value:
                    run_date = f"{date_value}T{time_value}:00"
                if not run_date:
                    raise HTTPException(status_code=400, detail="date is required for once schedule")
                scheduler.add_job(
                    run_fn,
                    "date",
                    id=job_key,
                    replace_existing=True,
                    run_date=run_date,
                    args=[job_id],
                )

            elif schedule_type == "monthly":
                day_of_month = int(data.get("date", 1))
                hour, minute = map(int, str(data.get("time", "")).split(":"))
                scheduler.add_job(
                    run_fn,
                    "cron",
                    id=job_key,
                    replace_existing=True,
                    day=day_of_month,
                    hour=hour,
                    minute=minute,
                    args=[job_id],
                )

            elif schedule_type == "cron":
                expr = str(data.get("cron", "")).strip()
                if not expr:
                    raise HTTPException(status_code=400, detail="cron is required for cron schedule")
                scheduler.add_job(
                    run_fn,
                    trigger=CronTrigger.from_crontab(expr),
                    id=job_key,
                    replace_existing=True,
                    args=[job_id],
                )
            else:
                raise HTTPException(status_code=400, detail="Unsupported schedule type")
        except HTTPException:
            raise
        except Exception as e:
            raise HTTPException(status_code=400, detail=f"Failed to schedule job: {str(e)}")

        return {"message": "Scheduled successfully", "action": action}


    @app.get("/schedules")
    def list_schedules():
        jobs = []
        for j in scheduler.get_jobs():
            if str(getattr(j, "id", "")).startswith("scheduled_job_"):
                jobs.append(_serialize_schedule_job(j))
        return {"items": jobs}


    @app.get("/schedules/{job_id}")
    def get_schedule(job_id: int, action: str | None = None):
        prefer_refresh = str(action or "").strip().lower() == "refresh"
        j, _key = _find_scheduled_job(job_id, prefer_refresh=prefer_refresh)
        if not j:
            raise HTTPException(status_code=404, detail="Schedule not found")
        return _serialize_schedule_job(j)


    @app.post("/schedules/{job_id}/pause")
    def pause_schedule(job_id: int, action: str | None = None):
        prefer_refresh = str(action or "").strip().lower() == "refresh"
        j, key = _find_scheduled_job(job_id, prefer_refresh=prefer_refresh)
        if not j:
            raise HTTPException(status_code=404, detail="Schedule not found")
        scheduler.pause_job(key)
        j = scheduler.get_job(key)
        return {"message": "Schedule paused", "schedule": _serialize_schedule_job(j)}


    @app.post("/schedules/{job_id}/resume")
    def resume_schedule(job_id: int, action: str | None = None):
        prefer_refresh = str(action or "").strip().lower() == "refresh"
        j, key = _find_scheduled_job(job_id, prefer_refresh=prefer_refresh)
        if not j:
            raise HTTPException(status_code=404, detail="Schedule not found")
        scheduler.resume_job(key)
        j = scheduler.get_job(key)
        return {"message": "Schedule resumed", "schedule": _serialize_schedule_job(j)}


    @app.delete("/schedules/{job_id}")
    def delete_schedule(job_id: int, action: str | None = None):
        prefer_refresh = str(action or "").strip().lower() == "refresh"
        j, key = _find_scheduled_job(job_id, prefer_refresh=prefer_refresh)
        if not j:
            raise HTTPException(status_code=404, detail="Schedule not found")
        scheduler.remove_job(key)
        return {"message": "Schedule deleted", "job_id": job_id}

    # --- SMART GETTERS (WITH STATS) ---

    # --- SMART GETTERS (WITH STATS) ---

    @app.get("/jobs")
    def get_all_jobs(db: Session = Depends(get_db)):
        # Keep newest jobs first so recent inserts are visible at the top in UI.
        jobs = db.query(models.Job).order_by(models.Job.job_id.desc()).all()
        result = []
        
        for job in jobs:
            tables = db.query(models.TableMetadata).filter(models.TableMetadata.job_id == job.job_id).all()
            
            total_rows = 0
            good_rows = 0
            error_rows = 0
            
            for t in tables:
                stat = db.query(models.TableStats).filter(
                    models.TableStats.job_id == job.job_id, 
                    models.TableStats.table_id == t.table_id
                ).order_by(models.TableStats.stat_id.desc()).first()
                
                if stat:
                    total_rows += (stat.total_rows or 0)
                    good_rows += (stat.good_rows or 0)
                    error_rows += ((stat.total_rows or 0) - (stat.good_rows or 0))
                else:
                    # Fallback for newly created jobs/tables before first run:
                    # show uploaded row_count from metadata instead of 0.
                    total_rows += (t.row_count or 0)

            covered_cols = db.query(distinct(models.Rule.column_name)).filter(models.Rule.job_id == job.job_id).count()

            # --- NEW: Calculate Job Duration in ms ---
            job_duration_str = "0ms"
            if getattr(job, 'start_time', None) and getattr(job, 'end_time', None):
                try:
                    duration_ms = (job.end_time - job.start_time).total_seconds() * 1000
                    job_duration_str = f"{duration_ms:.0f}ms"
                except Exception:
                    pass
            # -----------------------------------------

            result.append({
                "job_id": job.job_id,
                "job_name": job.job_name,
                "start_time": job.start_time.isoformat() if getattr(job, 'start_time', None) else None,
                "end_time": job.end_time.isoformat() if getattr(job, 'end_time', None) else None,
                "duration": job_duration_str, # <--- Now sending the ms duration
                "status": job.status,
                "total_tables": len(tables),
                "columns_covered": covered_cols, 
                "total_columns": db.query(models.ColumnMetadata).filter(models.ColumnMetadata.job_id == job.job_id).count(),
                "total_rules": db.query(models.Rule).filter(models.Rule.job_id == job.job_id, models.Rule.is_active == True).count(),
                "total_rows": total_rows,
                "good_rows": good_rows,
                "error_rows": error_rows
            })
        return result


    @app.get("/jobs/{job_id}/tables")
    def get_tables_for_job(job_id: int, db: Session = Depends(get_db)):
        tables = db.query(models.TableMetadata).filter(models.TableMetadata.job_id == job_id).all()
        result = []
        
        for t in tables:
            stat = db.query(models.TableStats).filter(
                models.TableStats.job_id == job_id, 
                models.TableStats.table_id == t.table_id
            ).order_by(models.TableStats.stat_id.desc()).first()
            
            rules_count = (
                db.query(models.Rule)
                .filter(
                    models.Rule.job_id == job_id,
                    models.Rule.table_id == t.table_id,
                    models.Rule.is_active == True,
                )
                .count()
            )
            
            col_count = db.query(models.ColumnMetadata).filter(
                models.ColumnMetadata.job_id == job_id, 
                models.ColumnMetadata.table_id == t.table_id
            ).count()
            
            # --- CHANGED: Calculate Table Duration in ms ---
            duration_str = "0ms"
            if stat and getattr(stat, 'end_time', None) and getattr(stat, 'start_time', None):
                try:
                    duration_ms = (stat.end_time - stat.start_time).total_seconds() * 1000
                    duration_str = f"{duration_ms:.0f}ms"
                except Exception:
                    pass 
            # -----------------------------------------------
            
            g_rows = stat.good_rows if stat and getattr(stat, 'good_rows', None) else 0
            v_errs = stat.validation_errors if stat and getattr(stat, 'validation_errors', None) else 0
            f_errs = stat.fuzzy_errors if stat and getattr(stat, 'fuzzy_errors', None) else 0
            t_rows = stat.total_rows if stat and getattr(stat, 'total_rows', None) else t.row_count
            
            result.append({
                "table_id": t.table_id,
                "table_name": t.table_name,
                "row_count": t_rows, 
                "column_count": col_count,
                "rule_count": rules_count,
                "good_rows": g_rows,
                "error_rows": (t_rows - g_rows),
                "duration": duration_str
            })
        return result 

    # In backend/main.py

    @app.get("/tables/{job_id}/{table_id}/details") # <--- CHANGED URL
    def get_table_details(job_id: int, table_id: int, db: Session = Depends(get_db)):
        # 1. Fetch specific table by BOTH Job ID and Table ID
        table = db.query(models.TableMetadata).filter(
            models.TableMetadata.job_id == job_id,
            models.TableMetadata.table_id == table_id
        ).first()

        if not table: return {"columns": [], "rules": []}
        
        # 2. Get Columns
        columns = db.query(models.ColumnMetadata).filter(
            models.ColumnMetadata.job_id == job_id,
            models.ColumnMetadata.table_id == table_id
        ).all()
        
        # 3. Get Rules
        rules = db.query(models.Rule).filter(
            models.Rule.job_id == job_id,
            models.Rule.table_id == table_id
        ).all()
        
        return {"columns": columns, "rules": rules}

    # --- RULE MANAGEMENT ---

    @app.post("/rules/add")
    def add_new_rule(rule: RuleCreate, db: Session = Depends(get_db)):
        new_rule = models.Rule(
            job_id=rule.job_id,
            table_id=rule.table_id,
            column_name=rule.column_name,
            rule_type=rule.rule_type,
            data_type=rule.data_type,
            rule_value=rule.rule_value,
            is_active=rule.is_active
        )
        db.add(new_rule)
        
        if rule.rule_type == "fuzzy_match" and rule.master_data:
            table = db.query(models.TableMetadata).filter(
                models.TableMetadata.job_id == rule.job_id,
                models.TableMetadata.table_id == rule.table_id
            ).first()
            if table:
                for val in rule.master_data:
                    db.add(models.MasterTable(
                        job_id=rule.job_id,
                        table_id=rule.table_id,
                        table_name=table.table_name,
                        master_value=val
                    ))
        db.commit()
        return {"message": "Rule Added"}

    @app.delete("/rules/{rule_id}")
    def delete_rule(rule_id: int, db: Session = Depends(get_db)):
        db.query(models.Rule).filter(models.Rule.rule_id == rule_id).delete()
        db.commit()
        return {"status": "deleted"}

    @app.put("/rules/{rule_id}/toggle")
    def toggle_rule(rule_id: int, payload: RuleToggle, db: Session = Depends(get_db)):
        rule = db.query(models.Rule).filter(models.Rule.rule_id == rule_id).first()
        if rule:
            rule.is_active = payload.is_active
            db.commit()
        return {"status": "updated"}

    @app.put("/rules/{rule_id}")
    def update_rule(rule_id: int, payload: RuleUpdate, db: Session = Depends(get_db)):
        # 1. Get the Rule
        rule = db.query(models.Rule).filter(models.Rule.rule_id == rule_id).first()
        if not rule:
            raise HTTPException(status_code=404, detail="Rule not found")
        
        # 2. Update Basic Fields
        rule.rule_type = payload.rule_type
        rule.rule_value = payload.rule_value
        rule.is_active = payload.is_active
        
        # 3. Handle Master Data Update (If Fuzzy Match)
        if payload.rule_type == "fuzzy_match":
            # Delete old master entries for this table
            db.query(models.MasterTable).filter(
                models.MasterTable.job_id == rule.job_id,
                models.MasterTable.table_id == rule.table_id
            ).delete()
            
            # Add new entries
            table = db.query(models.TableMetadata).filter(models.TableMetadata.table_id == rule.table_id).first()
            if table and payload.master_data:
                for val in payload.master_data:
                    db.add(models.MasterTable(
                        job_id=rule.job_id,
                        table_id=rule.table_id,
                        table_name=table.table_name,
                        master_value=val
                    ))
        
        db.commit()
        return {"message": "Rule Updated Successfully"}

    @app.get("/master-data/{job_id}/{table_id}")
    def get_master_data(job_id: int, table_id: int, db: Session = Depends(get_db)):
        """ Fetches the master list for a specific table (for editing) """
        masters = db.query(models.MasterTable).filter(
            models.MasterTable.job_id == job_id,
            models.MasterTable.table_id == table_id
        ).all()
        return [m.master_value for m in masters]

    @app.put("/jobs/{job_id}/rename")
    def rename_job(job_id: int, payload: RenamePayload, db: Session = Depends(get_db)):
        job = db.query(models.Job).filter(models.Job.job_id == job_id).first()
        if not job: raise HTTPException(status_code=404, detail="Job not found")
        job.job_name = payload.name
        db.commit()
        return {"message": "Job renamed successfully"}

    @app.put("/tables/{table_id}/rename")
    def rename_table(table_id: int, payload: dict, db: Session = Depends(get_db)):
        table = db.query(models.TableMetadata).filter(models.TableMetadata.table_id == table_id).first()
        if not table: 
            raise HTTPException(status_code=404, detail="Table not found")
        
        # FIX: Use .get() safely to find the name in the dictionary
        new_name = payload.get("name") or payload.get("new_name")
        
        if not new_name:
            raise HTTPException(status_code=400, detail="New name is required in payload")
        
        try:
            rename_table_csv(table.job_id, table.table_name, new_name)

            # --- UPDATE DATABASE ---
            table.table_name = new_name
            db.commit()
            return {"message": "Table renamed successfully"}
            
        except Exception as e:
            db.rollback()
            raise HTTPException(status_code=500, detail=f"Database error: {str(e)}")

    @app.delete("/jobs/{job_id}")
    def delete_job(job_id: int, db: Session = Depends(get_db)):
        job = db.query(models.Job).filter(models.Job.job_id == job_id).first()
        if not job: 
            raise HTTPException(status_code=404, detail="Job not found")
        
        try:
            # 1. Delete the "Leaf" nodes first (the newest tables we added)
            db.query(models.QuarantineLog).filter(models.QuarantineLog.job_id == job_id).delete()
            db.query(models.MasterTable).filter(models.MasterTable.job_id == job_id).delete()
            
            # 2. Delete the rest of the child records
            db.query(models.ColumnMetadata).filter(models.ColumnMetadata.job_id == job_id).delete()
            db.query(models.Rule).filter(models.Rule.job_id == job_id).delete()
            db.query(models.TableStats).filter(models.TableStats.job_id == job_id).delete()
            
            # 3. Delete the intermediate parent (Tables)
            db.query(models.TableMetadata).filter(models.TableMetadata.job_id == job_id).delete()
            
            # 4. Finally, it is safe to delete the root Job
            db.delete(job)
            db.commit()
            
            return {"message": "Job and all associated data deleted completely"}
            
        except Exception as e:
            db.rollback() # Instantly undo everything if it hits a snag
            raise HTTPException(status_code=500, detail=f"Failed to delete job: {str(e)}")

    @app.delete("/tables/{table_id}")
    def delete_table(table_id: int, db: Session = Depends(get_db)):
        table = db.query(models.TableMetadata).filter(models.TableMetadata.table_id == table_id).first()
        if not table: 
            raise HTTPException(status_code=404, detail="Table not found")
        
        try:
            # 1. Sweep away ALL associated child data first
            # Note: QuarantineLog tracks by job_id and table_name, so we use those here
            db.query(models.QuarantineLog).filter(
                models.QuarantineLog.job_id == table.job_id,
                models.QuarantineLog.table_name == table.table_name
            ).delete()
            
            db.query(models.MasterTable).filter(models.MasterTable.table_id == table_id).delete()
            db.query(models.ColumnMetadata).filter(models.ColumnMetadata.table_id == table_id).delete()
            db.query(models.Rule).filter(models.Rule.table_id == table_id).delete()
            db.query(models.TableStats).filter(models.TableStats.table_id == table_id).delete()
            
            # 2. Safely delete the parent table now that the children are gone
            db.delete(table)
            db.commit()
            return {"message": "Table and all associated data deleted completely"}
            
        except Exception as e:
            db.rollback() # Abort the transaction instantly if anything fails
            raise HTTPException(status_code=500, detail=f"Failed to delete table: {str(e)}")

    # --- DOWNLOAD ENDPOINTS (EXCEL WITH RED ERROR ROWS) ---

    def _load_result_dataframes(db: Session, table_name: str, job_id: int):
        """Load clean/error result tables for a job; fallback to latest table-name match."""
        inspector = inspect(db.bind)
        all_tables = inspector.get_table_names(schema="app_data")

        t_name_lower = table_name.lower()
        j_id_str = str(job_id)

        exact_clean = []
        exact_error = []
        fallback_clean = []
        fallback_error = []

        for t in all_tables:
            t_lower = t.lower()
            if t_name_lower not in t_lower:
                continue

            is_exact_job = (f"job{j_id_str}" in t_lower) or (f"_{j_id_str}_" in t_lower)
            if "clean" in t_lower:
                (exact_clean if is_exact_job else fallback_clean).append(t)
            elif "error" in t_lower or "quarantine" in t_lower or "bad" in t_lower:
                (exact_error if is_exact_job else fallback_error).append(t)

        # Prefer exact job match; if missing, use latest lexical match for same table name.
        actual_clean = sorted(exact_clean)[-1] if exact_clean else (sorted(fallback_clean)[-1] if fallback_clean else None)
        actual_error = sorted(exact_error)[-1] if exact_error else (sorted(fallback_error)[-1] if fallback_error else None)

        if not actual_clean and not actual_error:
            raise Exception(
                f"Tables missing in 'app_data'. Looked for '{t_name_lower}' + 'job{j_id_str}'. Exists in app_data: {all_tables}"
            )

        df_clean = pd.DataFrame()
        df_error = pd.DataFrame()
        if actual_clean:
            try:
                df_clean = pd.read_sql_table(actual_clean, db.bind, schema="app_data")
            except Exception:
                pass
        if actual_error:
            try:
                df_error = pd.read_sql_table(actual_error, db.bind, schema="app_data")
            except Exception:
                pass

        if df_clean.empty and df_error.empty:
            raise Exception(f"Found the tables ({actual_clean}, {actual_error}) in app_data, but they are completely empty.")
        return df_clean, df_error


    def generate_formatted_excel(db: Session, table_name: str, job_id: int):
        df_clean, df_error = _load_result_dataframes(db, table_name, job_id)

        # Combine and tag rows
        df_clean['__is_error__'] = False
        if not df_error.empty:
            df_error['__is_error__'] = True
            df_combined = pd.concat([df_clean, df_error], ignore_index=True)
        else:
            df_combined = df_clean
            
        # 5. Create Excel File
        output = io.BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df_combined.drop(columns=['__is_error__']).to_excel(writer, index=False, sheet_name='Data')
            worksheet = writer.sheets['Data']
            
            red_fill = PatternFill(start_color='FF9999', end_color='FF9999', fill_type='solid')
            
            for r_idx, is_error in enumerate(df_combined['__is_error__'], start=2):
                if is_error:
                    for cell in worksheet[r_idx]:
                        cell.fill = red_fill
                        
        output.seek(0)
        return output


    def generate_output_dataframe(db: Session, table_name: str, job_id: int):
        """Return combined clean+error rows as a dataframe."""
        df_clean, df_error = _load_result_dataframes(db, table_name, job_id)
        if not df_error.empty:
            return pd.concat([df_clean, df_error], ignore_index=True)
        return df_clean


    @app.get("/tables/{table_id}/download")
    def download_table_excel(table_id: int, db: Session = Depends(get_db)):
        table = db.query(models.TableMetadata).filter(models.TableMetadata.table_id == table_id).first()
        if not table:
            raise HTTPException(status_code=404, detail="Table not found")
            
        try:
            file_bytes, filename, mime_type = _build_table_output_bytes(db, table, "excel")
            response = StreamingResponse(iter([file_bytes]), media_type=mime_type)
            response.headers["Content-Disposition"] = f"attachment; filename={filename}"
            return response
        except Exception as e:
            raise HTTPException(status_code=404, detail=str(e))


    @app.get("/tables/{table_id}/download-csv")
    def download_table_csv(table_id: int, db: Session = Depends(get_db)):
        table = db.query(models.TableMetadata).filter(models.TableMetadata.table_id == table_id).first()
        if not table:
            raise HTTPException(status_code=404, detail="Table not found")
        try:
            file_bytes, filename, mime_type = _build_table_output_bytes(db, table, "csv")
            response = StreamingResponse(iter([file_bytes]), media_type=mime_type)
            response.headers["Content-Disposition"] = f"attachment; filename={filename}"
            return response
        except Exception as e:
            raise HTTPException(status_code=404, detail=str(e))


    def _resolve_table_for_job_and_table(db: Session, job_id: int, table_id: int):
        """Resolve table with a few safe fallbacks for mixed IDs from UI."""
        table = (
            db.query(models.TableMetadata)
            .filter(
                models.TableMetadata.job_id == job_id,
                models.TableMetadata.table_id == table_id,
            )
            .first()
        )
        if table:
            return table

        # Fallback 1: Some UI states can accidentally swap ids.
        swapped = (
            db.query(models.TableMetadata)
            .filter(
                models.TableMetadata.job_id == table_id,
                models.TableMetadata.table_id == job_id,
            )
            .first()
        )
        if swapped:
            return swapped

        # Fallback 2: If table_id is globally unique in current data, use latest match.
        return (
            db.query(models.TableMetadata)
            .filter(models.TableMetadata.table_id == table_id)
            .order_by(models.TableMetadata.job_id.desc())
            .first()
        )


    @app.get("/tables/{job_id}/{table_id}/download")
    def download_table_excel_by_job(job_id: int, table_id: int, db: Session = Depends(get_db)):
        table = _resolve_table_for_job_and_table(db, job_id, table_id)
        if not table:
            raise HTTPException(status_code=404, detail="Table not found for this job")
        try:
            file_bytes, filename, mime_type = _build_table_output_bytes(db, table, "excel")
            response = StreamingResponse(iter([file_bytes]), media_type=mime_type)
            response.headers["Content-Disposition"] = f"attachment; filename={filename}"
            return response
        except Exception as e:
            raise HTTPException(status_code=404, detail=str(e))


    @app.get("/tables/{job_id}/{table_id}/download-csv")
    def download_table_csv_by_job(job_id: int, table_id: int, db: Session = Depends(get_db)):
        table = _resolve_table_for_job_and_table(db, job_id, table_id)
        if not table:
            raise HTTPException(status_code=404, detail="Table not found for this job")
        try:
            file_bytes, filename, mime_type = _build_table_output_bytes(db, table, "csv")
            response = StreamingResponse(iter([file_bytes]), media_type=mime_type)
            response.headers["Content-Disposition"] = f"attachment; filename={filename}"
            return response
        except Exception as e:
            raise HTTPException(status_code=404, detail=str(e))


    def _build_table_output_bytes(db: Session, table, fmt: str):
        fmt = (fmt or "csv").strip().lower()
        source_csv_path = resolve_table_csv_path(table.job_id, table.table_name)

        if fmt == "excel":
            try:
                excel_io = generate_formatted_excel(db, table.table_name, table.job_id)
                return (
                    excel_io.getvalue(),
                    f"{table.table_name}_Results.xlsx",
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )
            except Exception:
                # Fallback for not-yet-run jobs: export raw uploaded CSV as Excel.
                if not source_csv_path:
                    raise
                df = pd.read_csv(source_csv_path)
                output = io.BytesIO()
                with pd.ExcelWriter(output, engine="openpyxl") as writer:
                    df.to_excel(writer, index=False, sheet_name="Data")
                output.seek(0)
                return (
                    output.getvalue(),
                    f"{table.table_name}_Source.xlsx",
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )

        if fmt != "csv":
            raise HTTPException(status_code=400, detail="format must be csv or excel")

        try:
            df = generate_output_dataframe(db, table.table_name, table.job_id)
            return (
                df.to_csv(index=False).encode("utf-8"),
                f"{table.table_name}_Results.csv",
                "text/csv",
            )
        except Exception:
            # Fallback for not-yet-run jobs: return raw uploaded CSV.
            if not source_csv_path:
                raise
            with open(source_csv_path, "rb") as f:
                return (f.read(), f"{table.table_name}_Source.csv", "text/csv")


    @app.post("/tables/{table_id}/email")
    def email_table_output(table_id: int, payload: TableEmailPayload, db: Session = Depends(get_db)):
        table = (
            db.query(models.TableMetadata)
            .filter(models.TableMetadata.table_id == table_id)
            .order_by(models.TableMetadata.job_id.desc())
            .first()
        )
        if not table:
            raise HTTPException(status_code=404, detail="Table not found")

        to_email = (payload.to_email or "").strip()
        if not to_email:
            raise HTTPException(status_code=400, detail="to_email is required")
        tenant_id = os.getenv("MS_GRAPH_TENANT_ID", "").strip()
        client_id = os.getenv("MS_GRAPH_CLIENT_ID", "").strip()
        client_secret = os.getenv("MS_GRAPH_CLIENT_SECRET", "").strip()
        sender_email = os.getenv("MS_GRAPH_SENDER_EMAIL", "").strip()

        if not tenant_id or not client_id or not client_secret or not sender_email:
            raise HTTPException(
                status_code=400,
                detail=(
                    "Email is not configured. Set MS_GRAPH_TENANT_ID, MS_GRAPH_CLIENT_ID, "
                    "MS_GRAPH_CLIENT_SECRET, and MS_GRAPH_SENDER_EMAIL in backend .env."
                ),
            )

        fmt = (payload.format or "csv").strip().lower()
        file_bytes, filename, mime_type = _build_table_output_bytes(db, table, fmt)

        subject = (
            payload.subject.strip()
            if payload.subject and payload.subject.strip()
            else f"MDQM Results - {table.table_name}"
        )
        body = (
            payload.body
            if payload.body is not None
            else (
                f"Hello,\n\n"
                f"Please find attached the MDQM output for table '{table.table_name}'.\n\n"
                f"Regards,\nMDQM"
            )
        )

        token_url = f"https://login.microsoftonline.com/{tenant_id}/oauth2/v2.0/token"
        token_body = urlparse.urlencode(
            {
                "client_id": client_id,
                "scope": "https://graph.microsoft.com/.default",
                "client_secret": client_secret,
                "grant_type": "client_credentials",
            }
        ).encode("utf-8")
        token_req = urlrequest.Request(
            token_url,
            data=token_body,
            headers={"Content-Type": "application/x-www-form-urlencoded"},
            method="POST",
        )

        try:
            with urlrequest.urlopen(token_req, timeout=20) as resp:
                token_payload = json.loads(resp.read().decode("utf-8"))
                access_token = token_payload.get("access_token")
        except Exception as e:
            raise HTTPException(status_code=400, detail=f"Failed to get Graph token: {str(e)}")

        if not access_token:
            raise HTTPException(status_code=400, detail="Graph token response missing access_token")

        graph_mail_url = f"https://graph.microsoft.com/v1.0/users/{sender_email}/sendMail"
        mail_payload = {
            "message": {
                "subject": subject,
                "body": {
                    "contentType": "HTML",
                    "content": body.replace("\n", "<br/>"),
                },
                "toRecipients": [
                    {
                        "emailAddress": {
                            "address": to_email,
                        }
                    }
                ],
                "attachments": [
                    {
                        "@odata.type": "#microsoft.graph.fileAttachment",
                        "name": filename,
                        "contentType": mime_type,
                        "contentBytes": base64.b64encode(file_bytes).decode("utf-8"),
                    }
                ],
            },
            "saveToSentItems": "true",
        }
        mail_req = urlrequest.Request(
            graph_mail_url,
            data=json.dumps(mail_payload).encode("utf-8"),
            headers={
                "Authorization": f"Bearer {access_token}",
                "Content-Type": "application/json",
            },
            method="POST",
        )

        try:
            with urlrequest.urlopen(mail_req, timeout=30):
                pass
        except Exception as e:
            raise HTTPException(status_code=400, detail=f"Failed to send email via Graph API: {str(e)}")

        return {
            "message": "Email sent successfully",
            "to_email": to_email,
            "filename": filename,
        }


    @app.post("/tables/{table_id}/sharepoint-upload")
    def upload_table_to_sharepoint(table_id: int, payload: dict = Body(...), db: Session = Depends(get_db)):
        """
        Upload table output file (csv/excel) to SharePoint via Microsoft Graph.
        Required env vars:
        SP_TENANT_ID, SP_CLIENT_ID, SP_CLIENT_SECRET, SP_DRIVE_ID
        Optional env vars:
        SP_FOLDER_PATH (default: MDQM-Exports)
        """
        table = (
            db.query(models.TableMetadata)
            .filter(models.TableMetadata.table_id == table_id)
            .order_by(models.TableMetadata.job_id.desc())
            .first()
        )
        if not table:
            raise HTTPException(status_code=404, detail="Table not found")

        tenant_id = os.getenv("SP_TENANT_ID")
        client_id = os.getenv("SP_CLIENT_ID")
        client_secret = os.getenv("SP_CLIENT_SECRET")
        drive_id = os.getenv("SP_DRIVE_ID")
        payload_folder = (payload.get("folder_path") or "").strip().strip("/")
        env_folder = (os.getenv("SP_FOLDER_PATH", "MDQM-Exports") or "MDQM-Exports").strip("/")
        folder_path = payload_folder or env_folder

        if not tenant_id or not client_id or not client_secret or not drive_id:
            raise HTTPException(
                status_code=400,
                detail="SharePoint is not configured. Set SP_TENANT_ID, SP_CLIENT_ID, SP_CLIENT_SECRET, SP_DRIVE_ID.",
            )

        fmt = (payload.get("format") or "csv").strip().lower()
        file_bytes, filename, _mime = _build_table_output_bytes(db, table, fmt)

        # 1) Acquire app-only token
        token_url = f"https://login.microsoftonline.com/{tenant_id}/oauth2/v2.0/token"
        token_body = urlparse.urlencode(
            {
                "grant_type": "client_credentials",
                "client_id": client_id,
                "client_secret": client_secret,
                "scope": "https://graph.microsoft.com/.default",
            }
        ).encode("utf-8")
        token_req = urlrequest.Request(
            token_url,
            data=token_body,
            headers={"Content-Type": "application/x-www-form-urlencoded"},
            method="POST",
        )

        try:
            with urlrequest.urlopen(token_req, timeout=20) as resp:
                token_payload = json.loads(resp.read().decode("utf-8"))
                access_token = token_payload.get("access_token")
        except Exception as e:
            raise HTTPException(status_code=400, detail=f"Failed to get SharePoint token: {str(e)}")

        if not access_token:
            raise HTTPException(status_code=400, detail="SharePoint token is missing access_token")

        # 2) Upload bytes to drive root:/folder/file:/content
        graph_path = f"{folder_path}/{filename}" if folder_path else filename
        upload_url = f"https://graph.microsoft.com/v1.0/drives/{drive_id}/root:/{graph_path}:/content"
        upload_req = urlrequest.Request(
            upload_url,
            data=file_bytes,
            headers={
                "Authorization": f"Bearer {access_token}",
                "Content-Type": "application/octet-stream",
            },
            method="PUT",
        )

        try:
            with urlrequest.urlopen(upload_req, timeout=60) as resp:
                item_payload = json.loads(resp.read().decode("utf-8"))
        except Exception as e:
            raise HTTPException(status_code=400, detail=f"Failed to upload to SharePoint: {str(e)}")

        return {
            "message": "Uploaded to SharePoint successfully",
            "filename": filename,
            "file_id": item_payload.get("id"),
            "web_url": item_payload.get("webUrl"),
        }


    @app.get("/jobs/{job_id}/download")
    def download_job_zip(job_id: int, db: Session = Depends(get_db)):
        tables = db.query(models.TableMetadata).filter(models.TableMetadata.job_id == job_id).all()
        
        zip_buffer = io.BytesIO()
        added_files = False
        error_logs = [] # Keep track of errors for debugging
        
        with zipfile.ZipFile(zip_buffer, "w", zipfile.ZIP_DEFLATED) as zip_file:
            for t in tables:
                try:
                    excel_io = generate_formatted_excel(db, t.table_name, job_id)
                    zip_file.writestr(f"{t.table_name}_Results.xlsx", excel_io.getvalue())
                    added_files = True
                except Exception as e:
                    error_logs.append(f"{t.table_name}: {str(e)}") 

        if not added_files:
            # If it fails now, it will tell you exactly WHY it failed in the browser error!
            raise HTTPException(status_code=404, detail=f"Download failed. Reasons: {error_logs}")

        zip_buffer.seek(0)
        response = StreamingResponse(iter([zip_buffer.getvalue()]), media_type="application/x-zip-compressed")
        response.headers["Content-Disposition"] = f"attachment; filename=Job_{job_id}_Results.zip"
        return response

    # --- QUARANTINE DASHBOARD ENDPOINTS ---

    @app.get("/quarantine/jobs")
    def get_quarantine_jobs(db: Session = Depends(get_db)):
        jobs = db.query(models.Job).all()
        result = []
        
        for job in jobs:
            tables = db.query(models.TableMetadata).filter(models.TableMetadata.job_id == job.job_id).all()
            
            val_errors_total = 0
            fuzzy_errors_total = 0
            
            for t in tables:
                stat = db.query(models.TableStats).filter(
                    models.TableStats.job_id == job.job_id, 
                    models.TableStats.table_id == t.table_id
                ).order_by(models.TableStats.stat_id.desc()).first()
                
                if stat:
                    val_errors_total += (stat.validation_errors or 0)
                    fuzzy_errors_total += (stat.fuzzy_errors or 0)

            # Only show jobs that actually have errors
            total_errors = val_errors_total + fuzzy_errors_total
            if total_errors > 0:
                result.append({
                    "job_id": job.job_id,
                    "job_name": job.job_name,
                    "total_tables": len(tables),
                    "total_errors": total_errors,
                    "validation_errors": val_errors_total,
                    "fuzzy_errors": fuzzy_errors_total
                })
                
        return result

    @app.get("/quarantine/jobs/{job_id}/tables")
    def get_quarantine_tables(job_id: int, db: Session = Depends(get_db)):
        tables = db.query(models.TableMetadata).filter(models.TableMetadata.job_id == job_id).all()
        result = []
        
        for t in tables:
            stat = db.query(models.TableStats).filter(
                models.TableStats.job_id == job_id, 
                models.TableStats.table_id == t.table_id
            ).order_by(models.TableStats.stat_id.desc()).first()
            
            col_count = db.query(models.ColumnMetadata).filter(
                models.ColumnMetadata.job_id == job_id, 
                models.ColumnMetadata.table_id == t.table_id
            ).count()
            
            v_errs = stat.validation_errors if stat and getattr(stat, 'validation_errors', None) else 0
            f_errs = stat.fuzzy_errors if stat and getattr(stat, 'fuzzy_errors', None) else 0
            t_rows = stat.total_rows if stat and getattr(stat, 'total_rows', None) else t.row_count
            
            # Only show tables that have errors
            if (v_errs + f_errs) > 0:
                result.append({
                    "table_id": t.table_id,
                    "table_name": t.table_name,
                    "total_rows": t_rows,
                    "total_columns": col_count,
                    "validation_errors": v_errs,
                    "fuzzy_errors": f_errs
                })
                
        return result

    # --- VALIDATION ERROR DETAILS ENDPOINTS ---

    @app.get("/quarantine/jobs/{job_id}/tables/{table_id}/validation")
    def get_validation_error_details(job_id: int, table_id: int, db: Session = Depends(get_db)):
        table_meta = db.query(models.TableMetadata).filter_by(job_id=job_id, table_id=table_id).first()
        if not table_meta: raise HTTPException(status_code=404, detail="Table not found")
        
        # 1. Get all column names and data types for this table
        cols = db.query(models.ColumnMetadata).filter_by(job_id=job_id, table_id=table_id).all()
        all_columns = [c.column_name for c in cols]
        col_types = {c.column_name: c.data_type for c in cols}
        
        # 2. Get the specific Validation Quarantine Logs
        logs = db.query(models.QuarantineLog).filter_by(
            job_id=job_id, table_name=table_meta.table_name, error_type="Validation"
        ).all()
        
        # 3. Read the full rows from the _error table so we can display all columns
        error_table = f"{table_meta.table_name}_job{job_id}_error".lower()
        df_records = []
        try:
            df = pd.read_sql_table(error_table, db.bind, schema="app_data")
            df_records = df.to_dict(orient="records")
        except Exception:
            pass
            
        results = []
        for log in logs:
            # Match the log to the full row data from the _error table
            matching_row = {}
            if df_records:
                for row in df_records:
                    if str(row.get(log.column_name)) == str(log.error_value):
                        matching_row = row
                        break
            
            results.append({
                "log_id": log.log_id,
                "error_column": log.column_name,
                "error_value": log.error_value,
                "data_type": col_types.get(log.column_name, "Unknown"),
                "description": log.description,
                "row_data": matching_row
            })
            
        return {
            "table_id": table_id,
            "table_name": table_meta.table_name,
            "total_errors": len(results),
            "all_columns": all_columns,
            "errors": results
        }

    @app.put("/quarantine/errors/{log_id}")
    def update_quarantine_error(log_id: int, payload: ErrorEdit, db: Session = Depends(get_db)):
        log = db.query(models.QuarantineLog).filter_by(log_id=log_id).first()
        if not log: raise HTTPException(status_code=404, detail="Log not found")
        
        # 1. Update the DB
        log.error_value = payload.new_value
        log.description = "Fixed Manually"
        
        # 2. Find the exact file
        file_path = resolve_table_csv_path(log.job_id, log.table_name)
        
        if not file_path:
            raise HTTPException(status_code=404, detail=f"Source CSV not found for job {log.job_id} table {log.table_name}")
            
        try:
            # Read the file
            df = pd.read_csv(file_path)
            
            row_index = int(log.row_id)
            
            if row_index >= len(df):
                raise Exception(f"Row {row_index} is out of bounds for file with {len(df)} rows.")
                
            if log.column_name not in df.columns:
                raise Exception(f"Column '{log.column_name}' not found in CSV.")
                
            # --- THE SMART TYPING FIX ---
            col_dtype = df[log.column_name].dtype
            new_val = payload.new_value
            
            try:
                # Dynamically cast the string to match the column's actual data type
                if pd.api.types.is_integer_dtype(col_dtype):
                    new_val = int(new_val)
                elif pd.api.types.is_float_dtype(col_dtype):
                    new_val = float(new_val)
                elif pd.api.types.is_bool_dtype(col_dtype):
                    new_val = str(new_val).lower() in ['true', '1', 'yes', 'y', 't']
            except ValueError:
                # If they try to type "abc" into an integer column, catch it and warn them!
                raise Exception(f"Invalid data type. Cannot save '{new_val}' into a numeric column.")
                
            # SURGERY: Overwrite the bad data safely
            df.loc[row_index, log.column_name] = new_val
            # -----------------------------
            
            # Save it back to the file
            df.to_csv(file_path, index=False)
            
        except Exception as e:
            raise HTTPException(status_code=500, detail=f"Failed to patch CSV: {str(e)}")

        db.commit()
        return {"message": "Error updated and source file patched successfully"}

    @app.delete("/quarantine/errors/{log_id}")
    def delete_quarantine_error(log_id: int, db: Session = Depends(get_db)):
        db.query(models.QuarantineLog).filter_by(log_id=log_id).delete()
        db.commit()
        return {"message": "Error deleted successfully"}

    # --- FUZZY ERROR DETAILS ENDPOINTS ---

    def _fuzzy_row_matches_tier(score: int, threshold: int, tier: str) -> bool:
        if tier == "below":
            return score < threshold
        if tier == "threshold-90":
            return threshold <= score <= 90
        if tier == "above-90":
            return score > 90
        return True


    @app.get("/quarantine/jobs/{job_id}/tables/{table_id}/fuzzy")
    def get_fuzzy_details(
        job_id: int,
        table_id: int,
        tier: str = "below",
        limit: int = 300,
        offset: int = 0,
        db: Session = Depends(get_db),
    ):
        """
        Paginated fuzzy analysis. Full-table JSON for 100k+ rows freezes browser and times out.
        Use tier + limit + offset. total_fuzzy_errors comes from latest TableStats when available.
        """
        if tier not in ("below", "threshold-90", "above-90"):
            raise HTTPException(status_code=400, detail="tier must be below, threshold-90, or above-90")
        limit = max(1, min(int(limit), 2000))
        offset = max(0, int(offset))

        table = db.query(models.TableMetadata).filter_by(job_id=job_id, table_id=table_id).first()
        if not table:
            raise HTTPException(status_code=404, detail="Table not found")

        rule = db.query(models.Rule).filter_by(job_id=job_id, table_id=table_id, rule_type="fuzzy_match").first()
        if not rule:
            raise HTTPException(status_code=404, detail="No fuzzy rule configured for this table.")

        threshold = int(rule.rule_value) if rule.rule_value else 70
        col_name = rule.column_name

        masters = db.query(models.MasterTable).filter_by(job_id=job_id, table_id=table_id).all()
        master_list = [m.master_value for m in masters]

        file_path = resolve_table_csv_path(job_id, table.table_name)
        if not file_path:
            raise HTTPException(status_code=404, detail="Source CSV not found for this job.")

        stat = (
            db.query(models.TableStats)
            .filter_by(job_id=job_id, table_id=table_id)
            .order_by(models.TableStats.stat_id.desc())
            .first()
        )
        total_fuzzy_errors = int(stat.fuzzy_errors or 0) if stat else 0

        df = pd.read_csv(file_path)
        all_columns = [str(c) for c in df.columns.tolist()]

        results = []
        skip_remaining = offset
        stopped_early = False

        for idx, row in df.iterrows():
            val = str(row.get(col_name, ""))
            if pd.isna(row.get(col_name)):
                val = ""

            best_match = "None"
            score = 0
            if master_list and val:
                match_tuple = process.extractOne(val, master_list)
                if match_tuple:
                    best_match = match_tuple[0]
                    score = int(match_tuple[1])

            is_error = score < threshold
            if not _fuzzy_row_matches_tier(score, threshold, tier):
                continue

            if skip_remaining > 0:
                skip_remaining -= 1
                continue

            if len(results) >= limit:
                stopped_early = True
                break

            def _cell_json(v):
                if v is None or (isinstance(v, float) and pd.isna(v)):
                    return ""
                if hasattr(v, "item"):
                    try:
                        return v.item()
                    except (ValueError, AttributeError):
                        return str(v)
                return v

            rd = row.fillna("")
            row_dict = {str(k): _cell_json(v) for k, v in rd.to_dict().items()}
            results.append(
                {
                    "row_id": int(idx),
                    "original_value": val,
                    "best_match": best_match,
                    "score": score,
                    "is_error": is_error,
                    "row_data": row_dict,
                }
            )

        has_more = stopped_early

        return {
            "table_name": table.table_name,
            "column_name": col_name,
            "threshold": threshold,
            "total_fuzzy_errors": total_fuzzy_errors,
            "master_list": master_list,
            "all_columns": all_columns,
            "data": results,
            "tier": tier,
            "limit": limit,
            "offset": offset,
            "has_more": has_more,
        }

    @app.post("/quarantine/jobs/{job_id}/tables/{table_id}/master")
    def add_to_master(job_id: int, table_id: int, payload: MasterAdd, db: Session = Depends(get_db)):
        table = db.query(models.TableMetadata).filter_by(table_id=table_id).first()
        # Check if it already exists to avoid duplicates
        exists = db.query(models.MasterTable).filter_by(job_id=job_id, table_id=table_id, master_value=payload.new_master).first()
        if not exists:
            db.add(models.MasterTable(job_id=job_id, table_id=table_id, table_name=table.table_name, master_value=payload.new_master))
            db.commit()
        return {"message": "Added to Master Data"}

    @app.put("/quarantine/jobs/{job_id}/tables/{table_id}/fuzzy/replace")
    def replace_fuzzy_value(job_id: int, table_id: int, payload: FuzzyReplace, db: Session = Depends(get_db)):
        table = db.query(models.TableMetadata).filter_by(table_id=table_id, job_id=job_id).first()
        if not table:
            raise HTTPException(status_code=404, detail="Table not found")
        file_path = resolve_table_csv_path(job_id, table.table_name)
        if not file_path:
            raise HTTPException(status_code=404, detail="Source CSV not found for this job.")
        
        try:
            df = pd.read_csv(file_path)
            # Bypass strict typing to safely inject the string alias
            df[payload.column_name] = df[payload.column_name].astype(object)
            df.loc[payload.row_id, payload.column_name] = payload.new_value
            df.to_csv(table_csv_path(job_id, table.table_name), index=False)
        except Exception as e:
            raise HTTPException(status_code=500, detail=str(e))
            
        return {"message": "Replaced in CSV"}

    # --- DASHBOARD SUMMARY ENDPOINT ---

    @app.get("/dashboard/summary")
    def get_dashboard_summary(db: Session = Depends(get_db)):
        # 1. System Overviews
        total_jobs = db.query(models.Job).count()
        total_tables = db.query(models.TableMetadata).count()
        total_rules = db.query(models.Rule).filter_by(is_active=True).count()

        # 2. Data Volume & Health (Calculated from the most recent run of every table)
        tables = db.query(models.TableMetadata).all()
        
        total_rows_processed = 0
        total_clean_rows = 0
        total_validation_errors = 0
        total_fuzzy_errors = 0

        for t in tables:
            # Get the absolute latest stat for this specific table
            latest_stat = db.query(models.TableStats).filter_by(table_id=t.table_id).order_by(models.TableStats.stat_id.desc()).first()
            
            if latest_stat:
                total_rows_processed += (latest_stat.total_rows or 0)
                total_clean_rows += (latest_stat.good_rows or 0)
                total_validation_errors += (latest_stat.validation_errors or 0)
                total_fuzzy_errors += (latest_stat.fuzzy_errors or 0)

        # 3. Calculate Overall Data Quality Score
        dq_score = 0.0
        if total_rows_processed > 0:
            dq_score = round((total_clean_rows / total_rows_processed) * 100, 2)

        return {
            "system_metrics": {
                "total_jobs": total_jobs,
                "total_tables": total_tables,
                "active_rules": total_rules
            },
            "data_health": {
                "overall_score": dq_score,
                "rows_processed": total_rows_processed,
                "clean_rows": total_clean_rows,
                "validation_errors": total_validation_errors,
                "fuzzy_errors": total_fuzzy_errors
            }
        }


    @app.get("/dashboard/data-quality-metrics")
    def get_data_quality_metrics(db: Session = Depends(get_db)):
        tables = db.query(models.TableMetadata).all()
        stats = db.query(models.TableStats).order_by(models.TableStats.stat_id.desc()).all()
        print("Raw TableStats:", stats)
        metrics = []
        all_rows_clean = True

        def _pct(numerator: float, denominator: float) -> float:
            if denominator <= 0:
                return 0.0
            return round((numerator / denominator) * 100, 2)

        for t in tables:
            latest_stat = (
                db.query(models.TableStats)
                .filter(
                    models.TableStats.job_id == t.job_id,
                    models.TableStats.table_id == t.table_id,
                )
                .order_by(models.TableStats.stat_id.desc())
                .first()
            )
            if not latest_stat:
                continue

            total_rows = int(latest_stat.total_rows or 0)
            good_rows = int(latest_stat.good_rows or 0)
            validation_errors = int(latest_stat.validation_errors or 0)
            fuzzy_errors = int(latest_stat.fuzzy_errors or 0)
            if total_rows == 0:
                print("Warning: total_rows is zero")
            if validation_errors > 0 or fuzzy_errors > 0 or good_rows < total_rows:
                all_rows_clean = False

            metrics.append(
                {
                    "job_id": t.job_id,
                    "table": t.table_name,
                    "completeness": _pct(good_rows, total_rows),
                    "accuracy": _pct(total_rows - fuzzy_errors, total_rows),
                    "consistency": _pct(total_rows - validation_errors, total_rows),
                    "uniqueness": _pct(total_rows - fuzzy_errors, total_rows),
                    "validity": _pct(total_rows - validation_errors, total_rows),
                    "timeliness": 95.0,
                }
            )

        if metrics and all_rows_clean:
            print("All rows are clean, no errors found")
        print("Computed Metrics:", metrics)

        return metrics
        
    @app.delete("/master-data/remove")
    def remove_master_value(payload: dict, db: Session = Depends(get_db)):
        jid = payload.get("job_id")
        tid = payload.get("table_id")
        val = payload.get("value")

        # Find the specific entry
        item = db.query(models.MasterTable).filter(
            models.MasterTable.job_id == jid,
            models.MasterTable.table_id == tid,
            models.MasterTable.master_value == val
        ).first()

        if not item:
            raise HTTPException(status_code=404, detail="Master value not found")

        db.delete(item)
        db.commit()
        return {"message": "Value removed from Master Data successfully"}


    @app.get("/tables/{table_id}/columns/stats")
    def get_table_column_stats(table_id: int, db: Session = Depends(get_db)):
        # 1. Get Table Metadata
        table = db.query(models.TableMetadata).filter(models.TableMetadata.table_id == table_id).first()
        if not table:
            raise HTTPException(status_code=404, detail="Table not found")

        # 2. Read the Physical CSV to get the current headers
        file_path = resolve_table_csv_path(table.job_id, table.table_name)
        if not file_path:
            return []

        df = pd.read_csv(file_path)
        all_columns = df.columns.tolist()
        total_rows = len(df)

        # 3. Query QuarantineLog for error counts per column
        # Group by column_name and count the IDs
        error_counts = db.query(
            models.QuarantineLog.column_name, 
            func.count(models.QuarantineLog.log_id)
        ).filter(
            models.QuarantineLog.job_id == table.job_id,
            models.QuarantineLog.table_name == table.table_name
        ).group_by(models.QuarantineLog.column_name).all()

        # Convert list of tuples [(col, count)] to a dictionary {col: count}
        error_map = {row[0]: row[1] for row in error_counts}

        # 4. Construct the final response
        stats = []
        for col in all_columns:
            errors = error_map.get(col, 0)
            stats.append({
                "column_name": col,
                "total": total_rows,
                "good": total_rows - errors,
                "errors": errors,
                "quality_pct": round(((total_rows - errors) / total_rows * 100), 2) if total_rows > 0 else 0
            })

        return stats

    @app.put("/tables/{table_id}/columns/rename")
    def rename_column(table_id: int, payload: dict, db: Session = Depends(get_db)):
        old_name = payload.get("old_name").strip()
        new_name = payload.get("new_name").strip()
        
        table = db.query(models.TableMetadata).filter(models.TableMetadata.table_id == table_id).first()
        if not table: raise HTTPException(status_code=404)

        try:
            # --- 1. DATABASE METADATA SYNC ---
            col_record = db.query(models.ColumnMetadata).filter(
                models.ColumnMetadata.job_id == table.job_id,
                models.ColumnMetadata.table_id == table_id,
                func.lower(models.ColumnMetadata.column_name) == old_name.lower()
            ).first()

            # Emergency Fallback if string mismatch
            if not col_record:
                all_cols = db.query(models.ColumnMetadata).filter(
                    models.ColumnMetadata.job_id == table.job_id,
                    models.ColumnMetadata.table_id == table_id
                ).all()
                col_record = next((c for c in all_cols if c.column_name.lower() in [old_name.lower(), "s. no.", "count", "list", "no"]), None)

            if col_record:
                actual_old_name = col_record.column_name
                col_record.column_name = new_name
                
                # Update Rules & Logs using the exact DB name
                db.query(models.Rule).filter(
                    models.Rule.job_id == table.job_id,
                    models.Rule.table_id == table_id,
                    models.Rule.column_name == actual_old_name
                ).update({"column_name": new_name}, synchronize_session=False)

                db.query(models.QuarantineLog).filter(
                    models.QuarantineLog.job_id == table.job_id,
                    models.QuarantineLog.column_name == actual_old_name
                ).update({"column_name": new_name}, synchronize_session=False)
                
                print(f"DB SYNC: {actual_old_name} -> {new_name}")

            # --- 2. PHYSICAL FILE SYNC ---
            file_path = resolve_table_csv_path(table.job_id, table.table_name)
            if file_path:
                df = pd.read_csv(file_path)
                
                # Identify the column in CSV (case-insensitive)
                csv_target = next((c for c in df.columns if c.lower() == old_name.lower()), None)
                
                if csv_target:
                    df.rename(columns={csv_target: new_name}, inplace=True)
                    
                    # REFINEMENT: Explicitly drop any garbage columns before saving
                    # This removes 'job_id', 'table_id', and any 'Unnamed' columns created by index=True
                    cols_to_keep = [
                        c for c in df.columns 
                        if c not in ['job_id', 'table_id'] 
                        and not c.startswith('Unnamed')
                    ]
                    
                    # Save only the legitimate data columns without index
                    scoped_path = table_csv_path(table.job_id, table.table_name)
                    df[cols_to_keep].to_csv(scoped_path, index=False)
                    print(f"FILE SYNC: {csv_target} -> {new_name} (Cleaned)")

            db.commit()
            return {"message": "Sync complete", "new_name": new_name}

        except Exception as e:
            db.rollback()
            print(f"CRITICAL ERROR: {e}")
            raise HTTPException(status_code=500, detail=str(e))
        
    @app.post("/tables/{table_id}/standardize-dates")
    def standardize_table_dates(table_id: int, payload: dict, db: Session = Depends(get_db)):
        column_name = payload.get("column_name")
        
        table = db.query(models.TableMetadata).filter(models.TableMetadata.table_id == table_id).first()
        rule = db.query(models.Rule).filter(
            models.Rule.table_id == table_id,
            models.Rule.column_name == column_name,
            models.Rule.rule_type == "date_format_check"
        ).first()

        if not rule:
            raise HTTPException(status_code=400, detail="No active date rule")

        try:
            file_path = resolve_table_csv_path(table.job_id, table.table_name)
            if not file_path:
                raise HTTPException(status_code=404, detail="Source CSV not found for this job")
            
            # 1. READ AS PURE TEXT (Prevents 00:00:00 on import)
            df = pd.read_csv(file_path, dtype=str, keep_default_na=False)

            def fix_date(val):
                if not val or pd.isna(val): return val
                
                # 2. CHOP OFF EXISTING TIMESTAMPS
                clean_val = str(val).split(" ")[0]
                
                # Clean separators
                clean_val = clean_val.replace("/", "-").replace(".", "-").replace("\\", "-")
                
                try:
                    # Try to parse the messy date
                    # We try a few formats if the primary one fails
                    for fmt in [rule.rule_value, "%Y-%m-%d", "%m-%d-%Y", "%d-%m-%y"]:
                        try:
                            date_obj = datetime.strptime(clean_val, fmt)
                            return date_obj.strftime(rule.rule_value)
                        except:
                            continue
                    return val
                except:
                    return val

            df[column_name] = df[column_name].apply(fix_date)

            # 3. SAVE WITHOUT INDEX (Prevents row jumbling/ID columns)
            # Only save columns that belong in the CSV
            original_cols = [c for c in df.columns if c not in ['job_id', 'table_id']]
            df[original_cols].to_csv(table_csv_path(table.job_id, table.table_name), index=False)

            return {"message": "Dates standardized successfully"}
        except Exception as e:
            print(f"Error: {e}")
            raise HTTPException(status_code=500, detail=str(e))
